"""
Etainabl API sync service for fetching and persisting site and supply data.
"""

import logging
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Set, Tuple
import requests
from openpyxl import load_workbook
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.utils import timezone as dj_timezone

from .api_client import EtainablApiClient
from .config_service import SettingsConfigService
from .models import (
    Site,
    Supply,
    AppSettings,
    AuditLogEntry,
    CapacityReference,
    CapacityUploadRun,
    ImportRun,
    HalfHourlyConsumption,
    MonthlyConsumption,
    InvoiceCost,
    MonthlyReport,
    MonthlyReportVersion,
    ReportComment,
    ReportPageValidationState,
    ReportValidationComment,
    ReportValidationEvent,
    ReportWriteGrant,
    ReportWriteDelegationEvent,
    Team,
    ReportOwnershipUnavailabilityApproval,
    ReportOwnershipTransferEvent,
)

logger = logging.getLogger(__name__)

CAPACITY_REQUIRED_HEADERS = (
    'Name',
    'eSight Meter Code',
    'Av Cap (kVA)',
)
CAPACITY_UPLOAD_STATUS_SUCCESS = CapacityUploadRun.STATUS_SUCCESS
CAPACITY_UPLOAD_STATUS_PARTIAL_SUCCESS = CapacityUploadRun.STATUS_PARTIAL_SUCCESS
CAPACITY_UPLOAD_STATUS_FAILED = CapacityUploadRun.STATUS_FAILED
CAPACITY_UPLOAD_ERROR_FILE_EMPTY = 'File is empty or missing a header row.'
CAPACITY_UPLOAD_ERROR_PARSE_PREFIX = 'Upload parsing failed:'
SITE_FLOOR_AREA_UNIT_SQM = 'sqm'
SITE_FLOOR_AREA_UNIT_SQFT = 'sqft'
COVER_SCOPE_TEMPLATE = (
    'This monthly energy report provides a consolidated overview of utility performance at {site_name}. '
    'It summarises electricity and water consumption using monthly invoice data, half-hourly electricity '
    'profiles, and daily usage comparisons. The report aims to highlight key trends, seasonal changes, and '
    'anomalies in consumption to support ongoing energy-performance management and cost-efficiency planning.'
)

# Audit log action constants.
AUDIT_ACTION_ADMIN_VIEW_AUDIT_LOG = 'ADMIN_VIEW_AUDIT_LOG'
AUDIT_ACTION_ADMIN_EXPORT_AUDIT_LOG = 'ADMIN_EXPORT_AUDIT_LOG'
AUDIT_ACTION_REPORT_SAVE_DRAFT = 'REPORT_SAVE_DRAFT'
AUDIT_ACTION_REPORT_SAVE_FINAL = 'REPORT_SAVE_FINAL'
AUDIT_ACTION_REPORT_REPLACE_FINAL = 'REPORT_REPLACE_FINAL'
AUDIT_ACTION_ADMIN_CREATE_INVITATION = 'ADMIN_CREATE_INVITATION'
AUDIT_ACTION_ADMIN_RESEND_INVITATION_EMAIL = 'ADMIN_RESEND_INVITATION_EMAIL'
AUDIT_ACTION_ADMIN_SEND_INVITATION_EMAIL = 'ADMIN_SEND_INVITATION_EMAIL'
AUDIT_ACTION_ADMIN_REVOKE_INVITATION = 'ADMIN_REVOKE_INVITATION'
AUDIT_ACTION_ADMIN_ACCEPT_INVITATION = 'ADMIN_ACCEPT_INVITATION'
AUDIT_ACTION_ADMIN_CREATE_TEAM = 'ADMIN_CREATE_TEAM'
AUDIT_ACTION_ADMIN_UPDATE_TEAM = 'ADMIN_UPDATE_TEAM'
AUDIT_ACTION_ADMIN_ADD_SUB_TEAM = 'ADMIN_ADD_SUB_TEAM'
AUDIT_ACTION_ADMIN_ASSIGN_TEAM = 'ADMIN_ASSIGN_TEAM'
AUDIT_ACTION_ADMIN_UNASSIGN_TEAM = 'ADMIN_UNASSIGN_TEAM'
AUDIT_ACTION_ADMIN_ASSIGN_ROLE = 'ADMIN_ASSIGN_ROLE'
AUDIT_ACTION_ADMIN_REVOKE_ROLE = 'ADMIN_REVOKE_ROLE'
AUDIT_ACTION_ADMIN_UPDATE_SETTINGS = 'ADMIN_UPDATE_SETTINGS'
AUDIT_ACTION_ADMIN_UPLOAD_CAPACITY = 'ADMIN_UPLOAD_CAPACITY'
AUDIT_ACTION_ADMIN_ENABLE_USER = 'ADMIN_ENABLE_USER'
AUDIT_ACTION_ADMIN_DISABLE_USER = 'ADMIN_DISABLE_USER'
AUDIT_ACTION_ADMIN_RESET_PASSWORD = 'ADMIN_RESET_PASSWORD'
AUDIT_ACTION_ADMIN_DELETE_USER = 'ADMIN_DELETE_USER'
AUDIT_ACTION_ADMIN_UPDATE_USERNAME = 'ADMIN_UPDATE_USERNAME'
AUDIT_ACTION_ADMIN_SYNC_TRIGGER = 'ADMIN_SYNC_TRIGGER'
AUDIT_ACTION_ACCESS_DENIED = 'ACCESS_DENIED'
AUDIT_ACTION_REPORT_GRANT_WRITE = 'REPORT_GRANT_WRITE'
AUDIT_ACTION_REPORT_REVOKE_WRITE = 'REPORT_REVOKE_WRITE'
AUDIT_ACTION_REPORT_TRANSFER_OWNERSHIP = 'REPORT_TRANSFER_OWNERSHIP'
AUDIT_ACTION_REPORT_APPROVE_UNAVAILABLE_OWNER = 'REPORT_APPROVE_UNAVAILABLE_OWNER'


def create_audit_log_entry(
    *,
    actor_user,
    action_type,
    action_outcome,
    target_entity_type,
    message,
    actor_username_snapshot='',
    source_ip=None,
    target_entity_id=None,
    target_entity_label=None,
    request_path='',
    metadata_json=None,
    retention_class='standard',
):
    """Persist one immutable audit row for administrative activity."""

    allowed_outcomes = {
        AuditLogEntry.OUTCOME_SUCCESS,
        AuditLogEntry.OUTCOME_DENIED,
        AuditLogEntry.OUTCOME_FAILED,
    }
    normalized_action_type = (action_type or '').strip()
    normalized_target_entity_type = (target_entity_type or '').strip()
    normalized_message = (message or '').strip()

    if not normalized_action_type:
        raise ValueError('action_type is required for audit logging')
    if action_outcome not in allowed_outcomes:
        raise ValueError(f'Unsupported action_outcome: {action_outcome}')
    if not normalized_target_entity_type:
        raise ValueError('target_entity_type is required for audit logging')
    if not normalized_message:
        raise ValueError('message is required for audit logging')

    username = (actor_username_snapshot or '').strip()
    if not username and actor_user is not None:
        username = actor_user.get_username()

    metadata_payload = metadata_json if isinstance(metadata_json, dict) else {}

    return AuditLogEntry.objects.create(
        actor_user=actor_user,
        actor_username_snapshot=username or 'unknown',
        source_ip=source_ip,
        action_type=normalized_action_type,
        action_outcome=action_outcome,
        target_entity_type=normalized_target_entity_type,
        target_entity_id=target_entity_id,
        target_entity_label=target_entity_label,
        message=normalized_message,
        request_path=request_path,
        metadata_json=metadata_payload,
        retention_class=retention_class,
    )


def get_filtered_audit_logs(*, filters):
    """Return audit queryset filtered with shared viewer/export semantics."""

    queryset = AuditLogEntry.objects.select_related('actor_user').order_by('-occurred_at_utc', '-created_at')

    user_obj = filters.get('user')
    if user_obj is not None:
        queryset = queryset.filter(actor_user=user_obj)

    keyword = (filters.get('keyword') or '').strip()
    if keyword:
        queryset = queryset.filter(
            Q(message__icontains=keyword)
            | Q(target_entity_label__icontains=keyword)
            | Q(target_entity_id__icontains=keyword)
            | Q(actor_username_snapshot__icontains=keyword)
        )

    start_dt = filters.get('start')
    if start_dt is not None:
        queryset = queryset.filter(occurred_at_utc__gte=start_dt)

    end_dt = filters.get('end')
    if end_dt is not None:
        queryset = queryset.filter(occurred_at_utc__lte=end_dt)

    action_type = (filters.get('action_type') or '').strip()
    if action_type:
        queryset = queryset.filter(action_type=action_type)

    return queryset


def check_audit_export_threshold(*, queryset, limit=50000):
    """Return (allowed, row_count) where allowed=False when export rows exceed limit."""

    row_count = queryset.count()
    return row_count <= limit, row_count


def serialize_audit_entry_for_export(entry):
    """Normalize one audit entry into tabular export fields."""

    return {
        'utc_timestamp': entry.occurred_at_utc.strftime('%Y-%m-%d %H:%M:%S UTC') if entry.occurred_at_utc else '',
        'actor_username': entry.actor_username_snapshot or '',
        'source_ip': entry.source_ip or '',
        'action_type': entry.action_type or '',
        'action_outcome': entry.action_outcome or '',
        'target_entity_type': entry.target_entity_type or '',
        'target_entity_id': entry.target_entity_id or '',
        'target_entity_label': entry.target_entity_label or '',
        'message': entry.message or '',
        'request_path': entry.request_path or '',
    }


@dataclass
class CoverContentsEntry:
    """Single line in front-cover-2 contents with optional meter suffix."""

    title: str
    meter_name: str = ''
    display_line: str = ''


@dataclass
class FrontCoverOneFields:
    """Editable fields and assets for front cover page 1."""

    site_title: str
    report_month_title: str
    report_date: str
    client_logo_asset: str = ''
    background_asset: str = '/static/sitesync/images/Green%20and%20Leafy%20Office.jpg'


@dataclass
class FrontCoverTwoFields:
    """Editable fields for front cover page 2."""

    scope_title: str = 'SCOPE'
    scope_body: str = ''
    contents_title: str = 'CONTENTS'
    contents_entries: List[CoverContentsEntry] = field(default_factory=list)


@dataclass
class ReportCoverSet:
    """Full cover package used by draft/final/PDF/PPTX variants."""

    report_context_id: str
    front_cover_1: FrontCoverOneFields
    front_cover_2: FrontCoverTwoFields
    back_cover: Dict[str, str]
    sequence: List[str] = field(default_factory=lambda: ['front_cover_1', 'front_cover_2', 'body_pages', 'back_cover'])


def format_cover_report_date(now_dt: Optional[datetime] = None) -> str:
    """Return fixed DD MMMM YYYY date string for cover rendering."""

    moment = now_dt or dj_timezone.now()
    return moment.strftime('%d %B %Y')


def build_scope_body_text(site_name: str) -> str:
    """Return canonical default scope text with site substitution."""

    resolved_site_name = (site_name or '').strip() or 'the selected site'
    return COVER_SCOPE_TEMPLATE.format(site_name=resolved_site_name)


def build_cover_contents_entries(supplies: Optional[List[Dict]]) -> List[CoverContentsEntry]:
    """Build default contents lines from visual titles and meter names."""

    entries: List[CoverContentsEntry] = [
        CoverContentsEntry(title='Total Utility Usage (\u00a3)', meter_name='', display_line='Total Utility Usage (\u00a3)'),
    ]

    for supply in supplies or []:
        utility = str(supply.get('utility_type_display') or supply.get('utility_type') or 'Utility').strip()
        utility_lower = utility.lower()
        meter_number = str(supply.get('meter_number') or '').strip()

        section_titles = [
            f'Monthly {utility_lower} usage overview',
            f'Monthly {utility_lower} consumption analysis',
        ]
        if utility_lower == 'electricity':
            section_titles.extend([
                'Electricity load factor and demand performance',
                'Half-hourly electricity usage comparison',
                'Daily electricity usage comparison \u2013 weekdays',
                'Daily electricity usage comparison \u2013 weekends',
            ])

        for title in section_titles:
            display_line = title
            if title != 'Total Utility Usage (\u00a3)' and meter_number:
                display_line = f'{title} ({meter_number})'
            entries.append(CoverContentsEntry(title=title, meter_name=meter_number, display_line=display_line))

    # Preserve order while de-duplicating repeated visual labels.
    deduped: List[CoverContentsEntry] = []
    seen_lines: Set[str] = set()
    for entry in entries:
        key = entry.display_line.casefold().strip()
        if key in seen_lines:
            continue
        seen_lines.add(key)
        deduped.append(entry)

    return deduped


def build_report_cover_set(site_name: str, end_month: str, supplies: Optional[List[Dict]] = None) -> Dict:
    """Build default report cover package for UI and export composition."""

    try:
        month_label = datetime.strptime(end_month, '%Y-%m').strftime('%B %Y')
    except (TypeError, ValueError):
        month_label = end_month or ''

    front_cover_1 = FrontCoverOneFields(
        site_title=(site_name or '').strip(),
        report_month_title=f'{month_label} Energy Report'.strip(),
        report_date=format_cover_report_date(),
    )

    front_cover_2 = FrontCoverTwoFields(
        scope_body=build_scope_body_text(site_name),
        contents_entries=build_cover_contents_entries(supplies),
    )

    cover_set = ReportCoverSet(
        report_context_id=f'{(site_name or '').strip()}::{(end_month or '').strip()}',
        front_cover_1=front_cover_1,
        front_cover_2=front_cover_2,
        back_cover={'image_asset': '/static/sitesync/images/Report%20Back%20Cover%20Page.jpg', 'is_editable': 'false'},
    )

    payload = asdict(cover_set)
    payload['front_cover_2']['contents_text'] = '\n'.join([entry['display_line'] for entry in payload['front_cover_2']['contents_entries']])
    return payload


def _build_capacity_upload_result(run: CapacityUploadRun) -> Dict:
    """Normalize upload-run persistence into the view-facing result contract."""
    return {
        'status': run.status,
        'total_rows': run.total_rows,
        'accepted_rows': run.accepted_rows,
        'rejected_rows': run.rejected_rows,
        'errors': run.error_summary,
        'run': run,
    }


def _normalize_floor_area_unit(value: Optional[str]) -> Optional[str]:
    """Normalize Etainabl floor-area unit values into sqm/sqft."""
    raw = str(value or '').strip().casefold()
    if not raw:
        return None
    if raw in {'sqm', 'sq m', 'sq metre', 'sq meter', 'square metre', 'square meter', 'm2', 'm^2', 'metric'}:
        return SITE_FLOOR_AREA_UNIT_SQM
    if raw in {'sqft', 'sq ft', 'square foot', 'square feet', 'ft2', 'ft^2', 'imperial'}:
        return SITE_FLOOR_AREA_UNIT_SQFT
    return None


def _extract_floor_area_from_value(value) -> Tuple[Optional[Decimal], Optional[str]]:
    """Extract floor-area value and unit from a direct or nested asset payload value."""
    if isinstance(value, dict):
        area_value = (
            value.get('value')
            or value.get('amount')
            or value.get('area')
            or value.get('size')
            or value.get('measurement')
        )
        unit_value = (
            value.get('unit')
            or value.get('uom')
            or value.get('unitType')
            or value.get('measurementUnit')
        )
        if area_value is None:
            return None, _normalize_floor_area_unit(unit_value)
        try:
            area = Decimal(str(area_value).strip())
        except (InvalidOperation, ValueError, AttributeError):
            return None, _normalize_floor_area_unit(unit_value)
        normalized_unit = _normalize_floor_area_unit(unit_value)
        if area == Decimal('1'):
            area = Decimal('0')
        return area, normalized_unit

    if value is None or str(value).strip() == '':
        return None, None
    try:
        area = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None, None
    if area == Decimal('1'):
        area = Decimal('0')
    return area, None


def _extract_site_floor_area(asset_data: Dict) -> Tuple[Optional[Decimal], Optional[str]]:
    """Extract site floor area and original unit from known or likely Etainabl asset keys."""
    candidates = [
        ('floorArea', 'floorAreaUnit'),
        ('floor_area', 'floor_area_unit'),
        ('grossInternalArea', 'grossInternalAreaUnit'),
        ('netInternalArea', 'netInternalAreaUnit'),
        ('internalArea', 'internalAreaUnit'),
        ('buildingArea', 'buildingAreaUnit'),
        ('area', 'areaUnit'),
    ]

    for area_key, unit_key in candidates:
        if area_key not in asset_data:
            continue
        area, unit = _extract_floor_area_from_value(asset_data.get(area_key))
        if unit is None:
            unit = _normalize_floor_area_unit(asset_data.get(unit_key))
        if area is not None or unit is not None:
            return area, unit

    for key, value in asset_data.items():
        if not isinstance(value, dict):
            continue
        if 'area' not in str(key).casefold() and 'floor' not in str(key).casefold():
            continue
        area, unit = _extract_floor_area_from_value(value)
        if area is not None or unit is not None:
            return area, unit

    return None, None


class EtainaibleSyncService:
    """Service to sync assets and accounts from Etainabl API."""
    
    def __init__(self):
        """Initialize the sync service with API configuration."""
        self.api_key = (settings.ETAINABL_API_KEY or '').strip()
        self.api_url = settings.ETAINABL_API_URL
        self.timeout = settings.API_TIMEOUT
        self.max_retries = 10
        self.base_backoff = 1  # Start with 1 second
        self.max_backoff = 120  # Cap at 120 seconds
        
        if not self.api_key:
            raise ValueError(
                "ETAINABL_API_KEY environment variable not set. "
                "Cannot initialize sync service."
            )
    
    def sync_all(self) -> Dict[str, int]:
        """
        Perform complete sync of assets and accounts.
        
        Returns:
            Dict with counts: {
                'sites_created': int,
                'sites_updated': int,
                'supplies_created': int,
                'supplies_updated': int,
            }
        """
        logger.info("Starting full sync of Etainabl data...")
        
        results = {
            'sites_created': 0,
            'sites_updated': 0,
            'sites_deleted': 0,
            'supplies_created': 0,
            'supplies_updated': 0,
            'supplies_deleted': 0,
        }
        
        try:
            # Sync assets (sites)
            sites_result = self.sync_assets()
            results['sites_created'] = sites_result.get('created', 0)
            results['sites_updated'] = sites_result.get('updated', 0)
            results['sites_deleted'] = sites_result.get('deleted', 0)
            logger.info(
                f"Assets sync complete: {results['sites_created']} created, "
                f"{results['sites_updated']} updated, "
                f"{results['sites_deleted']} deleted"
            )
            
            # Sync accounts (supplies)
            supplies_result = self.sync_accounts()
            results['supplies_created'] = supplies_result.get('created', 0)
            results['supplies_updated'] = supplies_result.get('updated', 0)
            results['supplies_deleted'] = supplies_result.get('deleted', 0)
            logger.info(
                f"Accounts sync complete: {results['supplies_created']} created, "
                f"{results['supplies_updated']} updated, "
                f"{results['supplies_deleted']} deleted"
            )
            
            logger.info(f"Full sync completed successfully: {results}")
            return results
            
        except Exception as e:
            logger.error(f"Sync failed with error: {str(e)}", exc_info=True)
            raise
    
    def sync_assets(self) -> Dict[str, int]:
        """
        Sync all assets (sites) from Etainabl.
        
        Returns:
            Dict with created and updated counts
        """
        logger.info("Starting asset sync...")
        created = 0
        updated = 0
        skipped = 0
        deleted = 0
        remote_site_ids: Set[str] = set()
        received_valid_payload = False
        
        try:
            endpoint = f"{self.api_url}/assets"
            page_size = 50
            page = 1
            
            while True:
                # Fetch paginated results
                params = {
                    'limit': page_size,
                    'page': page,
                }
                
                data = self._fetch_from_api(endpoint, params)
                
                assets = self._extract_data_items(data)
                if assets is None:
                    logger.warning(f"No data in response from {endpoint}")
                    break

                received_valid_payload = True

                if not assets:
                    logger.info(f"No more assets to fetch (page={page})")
                    break
                
                # Process each asset
                for asset in assets:
                    external_id = self._extract_site_external_id(asset)
                    if external_id:
                        remote_site_ids.add(external_id)
                    site_created = self._upsert_site(asset)
                    if site_created is True:
                        created += 1
                    elif site_created is False:
                        updated += 1
                    else:
                        skipped += 1
                
                # Check pagination
                total = data.get('total') if isinstance(data, dict) else None
                if total is not None:
                    reported_skip = data.get('skip') if isinstance(data, dict) else None
                    reported_limit = data.get('limit') if isinstance(data, dict) else None
                    offset = reported_skip if isinstance(reported_skip, int) else (page - 1) * page_size
                    limit_used = reported_limit if isinstance(reported_limit, int) else len(assets)
                    downloaded = offset + limit_used
                    if downloaded >= total:
                        logger.info(f"Reached end of assets (total={total})")
                        break

                if len(assets) < page_size:
                    logger.info("Reached final assets page by page size (page=%s)", page)
                    break

                page += 1

            if received_valid_payload:
                stale_sites = Site.objects.exclude(external_id__in=remote_site_ids)
                deleted = stale_sites.count()
                if deleted:
                    stale_sites.delete()
                    logger.info("Removed %s stale sites not present in Etainabl", deleted)
            else:
                logger.warning("Skipping stale site reconciliation due to invalid assets payload shape")
            
            logger.info(
                "Asset sync complete: %s created, %s updated, %s skipped, %s deleted",
                created,
                updated,
                skipped,
                deleted,
            )
            return {
                'created': created,
                'updated': updated,
                'skipped': skipped,
                'deleted': deleted,
            }
            
        except Exception as e:
            logger.error(f"Asset sync failed: {str(e)}", exc_info=True)
            raise
    
    def sync_accounts(self) -> Dict[str, int]:
        """
        Sync all accounts (supplies) from Etainabl.
        
        Returns:
            Dict with created and updated counts
        """
        logger.info("Starting account sync...")
        created = 0
        updated = 0
        skipped = 0
        deleted = 0
        remote_supply_ids: Set[str] = set()
        received_valid_payload = False
        
        try:
            endpoint = f"{self.api_url}/accounts"
            page_size = 50
            page = 1
            
            while True:
                # Fetch paginated results
                params = {
                    'limit': page_size,
                    'page': page,
                }
                
                data = self._fetch_from_api(endpoint, params)
                
                accounts = self._extract_data_items(data)
                if accounts is None:
                    logger.warning(f"No data in response from {endpoint}")
                    break

                received_valid_payload = True

                if not accounts:
                    logger.info(f"No more accounts to fetch (page={page})")
                    break
                
                # Process each account
                for account in accounts:
                    external_id = self._extract_supply_external_id(account)
                    if external_id:
                        remote_supply_ids.add(external_id)
                    supply_created = self._upsert_supply(account)
                    if supply_created is True:
                        created += 1
                    elif supply_created is False:
                        updated += 1
                    else:
                        skipped += 1
                
                # Check pagination
                total = data.get('total') if isinstance(data, dict) else None
                if total is not None:
                    reported_skip = data.get('skip') if isinstance(data, dict) else None
                    reported_limit = data.get('limit') if isinstance(data, dict) else None
                    offset = reported_skip if isinstance(reported_skip, int) else (page - 1) * page_size
                    limit_used = reported_limit if isinstance(reported_limit, int) else len(accounts)
                    downloaded = offset + limit_used
                    if downloaded >= total:
                        logger.info(f"Reached end of accounts (total={total})")
                        break

                if len(accounts) < page_size:
                    logger.info("Reached final accounts page by page size (page=%s)", page)
                    break

                page += 1

            if received_valid_payload:
                stale_supplies = Supply.objects.exclude(external_id__in=remote_supply_ids)
                deleted = stale_supplies.count()
                if deleted:
                    stale_supplies.delete()
                    logger.info("Removed %s stale supplies not present in Etainabl", deleted)
            else:
                logger.warning("Skipping stale supply reconciliation due to invalid accounts payload shape")
            
            logger.info(
                "Account sync complete: %s created, %s updated, %s skipped, %s deleted",
                created,
                updated,
                skipped,
                deleted,
            )
            return {
                'created': created,
                'updated': updated,
                'skipped': skipped,
                'deleted': deleted,
            }
            
        except Exception as e:
            logger.error(f"Account sync failed: {str(e)}", exc_info=True)
            raise
    
    def _fetch_from_api(self, endpoint: str, params: Optional[Dict] = None) -> Dict:
        """
        Fetch data from Etainabl API with retry logic.
        
        Args:
            endpoint: API endpoint URL
            params: Query parameters
        
        Returns:
            JSON response data
        
        Raises:
            Exception if all retries are exhausted
        """
        headers = {
            'x-api-key': self.api_key,
            'Authorization': f'Bearer {self.api_key}',
            'Content-Type': 'application/json',
        }
        
        for attempt in range(1, self.max_retries + 1):
            try:
                logger.debug(
                    f"API request to {endpoint} (attempt {attempt}/{self.max_retries})"
                )
                
                response = requests.get(
                    endpoint,
                    params=params,
                    headers=headers,
                    timeout=self.timeout,
                )
                
                # Check for permanent errors
                if response.status_code in [400, 401, 403, 404]:
                    logger.error(
                        f"Permanent API error {response.status_code}: {response.text}"
                    )
                    raise Exception(
                        f"API returned {response.status_code}: {response.text}"
                    )
                
                # Retry on transient errors
                if response.status_code >= 500 or response.status_code == 408:
                    if attempt < self.max_retries:
                        backoff = min(
                            self.base_backoff * (2 ** (attempt - 1)),
                            self.max_backoff
                        )
                        logger.warning(
                            f"Transient API error {response.status_code}, "
                            f"retrying in {backoff}s..."
                        )
                        time.sleep(backoff)
                        continue
                    else:
                        raise Exception(
                            f"API returned {response.status_code} after "
                            f"{self.max_retries} retries"
                        )
                
                # Success
                response.raise_for_status()
                return response.json()
                
            except requests.exceptions.Timeout:
                if attempt < self.max_retries:
                    backoff = min(
                        self.base_backoff * (2 ** (attempt - 1)),
                        self.max_backoff
                    )
                    logger.warning(f"Timeout, retrying in {backoff}s...")
                    time.sleep(backoff)
                    continue
                else:
                    logger.error(f"Timeout after {self.max_retries} retries")
                    raise
            
            except requests.exceptions.RequestException as e:
                if attempt < self.max_retries:
                    backoff = min(
                        self.base_backoff * (2 ** (attempt - 1)),
                        self.max_backoff
                    )
                    logger.warning(f"Request error: {str(e)}, retrying in {backoff}s...")
                    time.sleep(backoff)
                    continue
                else:
                    logger.error(f"Request failed after {self.max_retries} retries: {str(e)}")
                    raise
        
        raise Exception(f"Failed to fetch from {endpoint} after {self.max_retries} attempts")

    def _extract_data_items(self, payload: object) -> Optional[List[Dict]]:
        """Extract list-like data from known Etainabl response shapes."""
        if payload is None:
            return None
        if isinstance(payload, list):
            return payload
        if not isinstance(payload, dict):
            return None

        for key in ('data', 'results', 'items'):
            value = payload.get(key)
            if isinstance(value, list):
                return value

        return None
    
    def _upsert_site(self, asset_data: Dict) -> Optional[bool]:
        """
        Create or update a Site from asset data.
        
        Args:
            asset_data: Asset data from Etainabl API
        
        Returns:
            True if created, False if updated
        """
        try:
            external_id = self._extract_site_external_id(asset_data)
            if not external_id:
                logger.warning(f"Asset missing id field: {asset_data}")
                return None

            name_value = (
                asset_data.get('name')
                or asset_data.get('siteName')
                or asset_data.get('siteCode')
                or asset_data.get('label')
            )
            if isinstance(name_value, dict):
                name_value = (
                    name_value.get('name')
                    or name_value.get('value')
                    or name_value.get('label')
                )

            description_value = asset_data.get('description')
            if not description_value:
                address_data = asset_data.get('address')
                if isinstance(address_data, dict):
                    address_parts = [
                        address_data.get('streetAddress'),
                        address_data.get('locality'),
                        address_data.get('region'),
                        address_data.get('postCode'),
                    ]
                    description_value = ", ".join(
                        str(part).strip() for part in address_parts if part
                    )
                elif address_data:
                    description_value = str(address_data)

            if isinstance(description_value, dict):
                description_value = str(description_value)
            floor_area_value, floor_area_unit = _extract_site_floor_area(asset_data)
            
            site, created = Site.objects.update_or_create(
                external_id=external_id,
                defaults={
                    'name': str(name_value or 'Unknown'),
                    'description': str(description_value or ''),
                    'floor_area': floor_area_value,
                    'floor_area_unit': floor_area_unit,
                }
            )
            
            if created:
                logger.debug(f"Created site: {site.name} (id={external_id})")
            else:
                logger.debug(f"Updated site: {site.name} (id={external_id})")
            
            return created
            
        except Exception as e:
            logger.error(f"Failed to upsert site from {asset_data}: {str(e)}")
            return None
    
    def _upsert_supply(self, account_data: Dict) -> Optional[bool]:
        """
        Create or update a Supply from account data.
        
        Args:
            account_data: Account data from Etainabl API
        
        Returns:
            True if created, False if updated
        """
        try:
            external_id = self._extract_supply_external_id(account_data)
            site_external_id = (
                account_data.get('asset_id')
                or account_data.get('assetId')
                or account_data.get('asset')
            )
            parent_account_id = (
                account_data.get('parentAccountId')
                or account_data.get('parent_account_id')
                or account_data.get('parentAccount')
            )
            if isinstance(site_external_id, dict):
                site_external_id = (
                    site_external_id.get('id')
                    or site_external_id.get('_id')
                    or site_external_id.get('assetId')
                )
            if isinstance(parent_account_id, dict):
                parent_account_id = (
                    parent_account_id.get('id')
                    or parent_account_id.get('_id')
                    or parent_account_id.get('accountId')
                )
            
            if not external_id or not site_external_id:
                logger.warning(f"Account missing id or asset_id: {account_data}")
                return None
            
            # Find associated site
            try:
                site = Site.objects.get(external_id=site_external_id)
            except Site.DoesNotExist:
                logger.warning(
                    f"Site not found for account {external_id} "
                    f"(site_id={site_external_id})"
                )
                return None
            
            # Map utility type
            utility_type_map = {
                'electricity': 'electricity',
                'electric': 'electricity',
                'gas': 'gas',
                'water': 'water',
                'thermal': 'other',
            }
            utility_raw = str(account_data.get('type', account_data.get('utility_type', 'other'))).lower()
            utility_type = utility_type_map.get(utility_raw, 'other')
            status = self._normalize_supply_status(account_data)

            device_id = account_data.get('device_id') or account_data.get('deviceId') or ''
            if not device_id:
                third_parties = account_data.get('thirdParties') or []
                if isinstance(third_parties, list) and third_parties:
                    first_tp = third_parties[0]
                    if isinstance(first_tp, dict):
                        device_id = first_tp.get('deviceId') or ''
            
            supply, created = Supply.objects.update_or_create(
                external_id=external_id,
                defaults={
                    'site': site,
                    'name': str(account_data.get('name') or account_data.get('label') or 'Unknown'),
                    'utility_type': utility_type,
                    'device_id': device_id,
                    'parent_account_id': str(parent_account_id) if parent_account_id else None,
                    'status': status,
                }
            )
            
            if created:
                logger.debug(f"Created supply: {supply.name} (id={external_id})")
            else:
                logger.debug(f"Updated supply: {supply.name} (id={external_id})")
            
            return created
            
        except Exception as e:
            logger.error(f"Failed to upsert supply from {account_data}: {str(e)}")
            return None

    def _extract_site_external_id(self, asset_data: Dict) -> Optional[str]:
        """Extract a stable site external ID from asset payload."""
        external_id = (
            asset_data.get('id')
            or asset_data.get('_id')
            or asset_data.get('assetId')
        )
        if not external_id:
            return None
        return str(external_id).strip()

    def _extract_supply_external_id(self, account_data: Dict) -> Optional[str]:
        """Extract a stable supply external ID from account payload."""
        external_id = account_data.get('id') or account_data.get('_id')
        if not external_id:
            return None
        return str(external_id).strip()

    def _normalize_supply_status(self, account_data: Dict) -> Optional[str]:
        """Extract and normalize account status from Etainabl payload."""
        raw_status = (
            account_data.get('status')
            or account_data.get('state')
            or account_data.get('accountStatus')
        )
        if isinstance(raw_status, dict):
            raw_status = (
                raw_status.get('status')
                or raw_status.get('state')
                or raw_status.get('value')
                or raw_status.get('label')
            )
        if raw_status is None:
            return None

        status = str(raw_status).strip().lower()
        return status or None


def parse_utc_datetime(value: str) -> datetime:
    """Parse ISO datetime/date values from source and return UTC-aware datetime."""
    if not value:
        raise ValueError('Missing datetime value')

    normalized = value.replace('Z', '+00:00')
    parsed = datetime.fromisoformat(normalized)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def canonical_month_key(value: datetime) -> str:
    """Derive canonical month key from UTC datetime.

    Etainabl uses exclusive end dates: a period ending on 2026-06-01T00:00:00Z
    covers up to (but not including) June 1st, so it belongs to May 2026.
    When the timestamp is exactly midnight on the 1st of a month we subtract one
    second to land in the correct (previous) month.
    """
    dt = value.astimezone(timezone.utc)
    if dt.day == 1 and dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        dt = dt - timedelta(seconds=1)
    return dt.strftime('%Y-%m')


def month_start(year: int, month: int) -> datetime:
    return datetime(year, month, 1, tzinfo=timezone.utc)


def shift_months(start: datetime, months: int) -> datetime:
    total = start.year * 12 + (start.month - 1) + months
    year = total // 12
    month = (total % 12) + 1
    return datetime(year, month, 1, tzinfo=timezone.utc)


def reporting_month_bounds(reporting_month: str) -> Tuple[datetime, datetime]:
    year, month = reporting_month.split('-')
    start = month_start(int(year), int(month))
    end = shift_months(start, 1)
    return start, end


def format_api_datetime(value: datetime) -> str:
    """Format UTC datetime for Xcelerate API requests (ISO with millisecond precision)."""
    return value.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')


def get_halfhourly_windows(reporting_month: str) -> List[Tuple[datetime, datetime]]:
    start, end = reporting_month_bounds(reporting_month)
    prior_start = shift_months(start, -12)
    prior_end = shift_months(end, -12)
    return [(start, end), (prior_start, prior_end)]


def get_monthly_window(reporting_month: str) -> Tuple[datetime, datetime]:
    start, end = reporting_month_bounds(reporting_month)
    return shift_months(end, -settings.CONSUMPTION_MONTHLY_MONTHS), end


def get_invoice_window(reporting_month: str) -> Tuple[datetime, datetime]:
    start, end = reporting_month_bounds(reporting_month)
    return shift_months(end, -settings.CONSUMPTION_INVOICE_MONTHS), end


def normalize_reporting_month(end_month: Optional[str], reporting_month: Optional[str]) -> str:
    """Normalize dual parameter naming to one canonical report month key."""
    candidate = (end_month or '').strip() or (reporting_month or '').strip()
    if not candidate:
        now = dj_timezone.localtime(dj_timezone.now())
        candidate = shift_months(month_start(now.year, now.month), -1).strftime('%Y-%m')
    # Reuse existing bounds parser as format validation.
    reporting_month_bounds(candidate)
    return candidate


def normalize_capacity_header(value: Optional[str]) -> str:
    """Normalize upload headers with trim + casefold rules."""
    return str(value or '').strip().casefold()


def normalize_esight_meter_code(value: Optional[str]) -> str:
    """Normalize eSight meter code for deterministic matching and storage."""
    return str(value or '').strip().upper()


def get_capacity_lookup_by_meter_codes(meter_codes: List[str]) -> Dict[str, Decimal]:
    """Return normalized meter-code -> available_capacity_kva lookup map."""
    normalized = [normalize_esight_meter_code(code) for code in meter_codes if normalize_esight_meter_code(code)]
    if not normalized:
        return {}
    return {
        row.esight_meter_code: row.available_capacity_kva
        for row in CapacityReference.objects.filter(esight_meter_code__in=normalized)
    }


def import_capacity_upload(uploaded_file) -> Dict:
    """Import available capacity reference rows from an uploaded .xlsx file."""
    filename = getattr(uploaded_file, 'name', '')
    required_headers_normalized = {
        normalize_capacity_header(header): header
        for header in CAPACITY_REQUIRED_HEADERS
    }
    row_errors = []
    accepted_rows = 0
    total_rows = 0

    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)

        if header_row is None:
            run = CapacityUploadRun.objects.create(
                uploaded_filename=filename,
                total_rows=0,
                accepted_rows=0,
                rejected_rows=0,
                status=CAPACITY_UPLOAD_STATUS_FAILED,
                error_summary=[CAPACITY_UPLOAD_ERROR_FILE_EMPTY],
            )
            return _build_capacity_upload_result(run)

        header_values = [normalize_capacity_header(cell) for cell in header_row]
        missing = [
            canonical
            for normalized, canonical in required_headers_normalized.items()
            if normalized not in header_values
        ]
        if missing:
            run = CapacityUploadRun.objects.create(
                uploaded_filename=filename,
                total_rows=0,
                accepted_rows=0,
                rejected_rows=0,
                status=CAPACITY_UPLOAD_STATUS_FAILED,
                error_summary=[f"Missing required columns: {', '.join(missing)}"],
            )
            return _build_capacity_upload_result(run)

        indexes = {
            canonical: header_values.index(normalized)
            for normalized, canonical in required_headers_normalized.items()
        }

        seen_codes = set()
        valid_rows = []
        for row_index, row in enumerate(row_iter, start=2):
            if row is None:
                continue
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            total_rows += 1
            name_raw = row[indexes['Name']] if len(row) > indexes['Name'] else None
            code_raw = row[indexes['eSight Meter Code']] if len(row) > indexes['eSight Meter Code'] else None
            capacity_raw = row[indexes['Av Cap (kVA)']] if len(row) > indexes['Av Cap (kVA)'] else None

            name = str(name_raw or '').strip()
            code = normalize_esight_meter_code(code_raw)
            current_errors = []

            if not name:
                current_errors.append('Name is blank')
            if not code:
                current_errors.append('eSight Meter Code is blank')

            capacity_value = None
            capacity_text = '' if capacity_raw is None else str(capacity_raw).strip()
            if not capacity_text:
                current_errors.append('Av Cap (kVA) is blank')
            else:
                try:
                    capacity_value = Decimal(capacity_text)
                    if capacity_value < 0:
                        current_errors.append('Av Cap (kVA) cannot be negative')
                except (InvalidOperation, ValueError):
                    current_errors.append('Av Cap (kVA) must be numeric when provided')

            if code and code in seen_codes:
                current_errors.append('Duplicate eSight Meter Code in upload')

            if current_errors:
                row_errors.append(f"Row {row_index}: {', '.join(current_errors)}")
                continue

            seen_codes.add(code)
            valid_rows.append({
                'name': name,
                'esight_meter_code': code,
                'available_capacity_kva': capacity_value,
            })

        with transaction.atomic():
            now = dj_timezone.now()
            for valid_row in valid_rows:
                CapacityReference.objects.update_or_create(
                    esight_meter_code=valid_row['esight_meter_code'],
                    defaults={
                        'name': valid_row['name'],
                        'available_capacity_kva': valid_row['available_capacity_kva'],
                        'source_filename': filename,
                        'last_imported_at': now,
                    },
                )
                accepted_rows += 1

            rejected_rows = len(row_errors)
            if accepted_rows > 0 and rejected_rows == 0:
                status = CAPACITY_UPLOAD_STATUS_SUCCESS
            elif accepted_rows > 0 and rejected_rows > 0:
                status = CAPACITY_UPLOAD_STATUS_PARTIAL_SUCCESS
            else:
                status = CAPACITY_UPLOAD_STATUS_FAILED

            run = CapacityUploadRun.objects.create(
                uploaded_filename=filename,
                total_rows=total_rows,
                accepted_rows=accepted_rows,
                rejected_rows=rejected_rows,
                status=status,
                error_summary=row_errors,
            )

        return _build_capacity_upload_result(run)
    except Exception as exc:  # pylint: disable=broad-except
        run = CapacityUploadRun.objects.create(
            uploaded_filename=filename,
            total_rows=0,
            accepted_rows=0,
            rejected_rows=0,
            status=CAPACITY_UPLOAD_STATUS_FAILED,
            error_summary=[f'{CAPACITY_UPLOAD_ERROR_PARSE_PREFIX} {str(exc)}'],
        )
        return _build_capacity_upload_result(run)


def get_or_create_monthly_report(site: Site, reporting_month: str, actor_user=None) -> MonthlyReport:
    """Return the unique monthly report identity for site + reporting month."""
    defaults = {'current_status': MonthlyReport.STATUS_DRAFT}
    if actor_user is not None:
        defaults['owner_user'] = actor_user
        defaults['created_by_user'] = actor_user
        defaults['last_modified_by_user'] = actor_user
        defaults['last_modified_at'] = dj_timezone.now()

    report, _ = MonthlyReport.objects.get_or_create(
        site=site,
        reporting_month=reporting_month,
        defaults=defaults,
    )

    if report.owner_user_id is None and actor_user is not None:
        report.owner_user = actor_user
        if report.created_by_user_id is None:
            report.created_by_user = actor_user
        if report.last_modified_by_user_id is None:
            report.last_modified_by_user = actor_user
            report.last_modified_at = dj_timezone.now()
        report.save(update_fields=['owner_user', 'created_by_user', 'last_modified_by_user', 'last_modified_at', 'updated_at'])

    return report


def get_report_validation_page_keys(report: MonthlyReport) -> List[str]:
    """Return canonical page keys for validation state tracking."""
    version = report.current_version
    if version is None and report.current_final_version is not None:
        version = report.current_final_version
    if version is None:
        return []
    return list(version.comments.order_by('visual_key').values_list('visual_key', flat=True))


def ensure_report_validation_rows(report: MonthlyReport) -> List[ReportPageValidationState]:
    """Ensure every canonical page key has a validation state row."""
    existing_qs = ReportPageValidationState.objects.filter(report=report)
    page_keys = get_report_validation_page_keys(report)
    if not page_keys:
        return list(existing_qs.order_by('page_key'))

    existing_rows = {
        row.page_key: row
        for row in existing_qs.filter(page_key__in=page_keys)
    }
    rows = []
    for page_key in page_keys:
        row = existing_rows.get(page_key)
        if row is None:
            row = ReportPageValidationState.objects.create(report=report, page_key=page_key)
        rows.append(row)
    return rows


def ensure_report_validation_rows_for_keys(report: MonthlyReport, page_keys: List[str]) -> List[ReportPageValidationState]:
    """Ensure validation rows exist for the provided page keys."""
    normalized_keys = []
    for key in page_keys or []:
        normalized = str(key or '').strip()
        if normalized:
            normalized_keys.append(normalized)

    if not normalized_keys:
        return ensure_report_validation_rows(report)

    unique_keys = list(dict.fromkeys(normalized_keys))
    existing_rows = {
        row.page_key: row
        for row in ReportPageValidationState.objects.filter(report=report, page_key__in=unique_keys)
    }
    rows = []
    for page_key in unique_keys:
        row = existing_rows.get(page_key)
        if row is None:
            row = ReportPageValidationState.objects.create(report=report, page_key=page_key)
        rows.append(row)
    return rows


def get_report_validation_summary(report: MonthlyReport) -> Dict[str, object]:
    """Return a report-level validation summary for editor and listing views."""
    rows = ensure_report_validation_rows(report)
    total_page_count = len(rows)
    validated_page_count = sum(1 for row in rows if row.is_validated)
    can_finalize = (
        report.validation_status == MonthlyReport.VALIDATION_VALIDATED
        and total_page_count > 0
        and validated_page_count == total_page_count
    )
    pages_validation = {
        row.page_key: {
            'page_key': row.page_key,
            'is_validated': row.is_validated,
            'validated_by_user_id': row.validated_by_user_id,
            'validated_by_user_name': row.validated_by_user.get_username() if row.validated_by_user else None,
            'validated_at': row.validated_at,
            'reset_reason': row.reset_reason,
            'reset_at': row.reset_at,
        }
        for row in rows
    }
    return {
        'validation_status': report.validation_status,
        'validator_user_id': report.validator_user_id,
        'validator_user': report.validator_user,
        'validator_assigned_by_user_id': report.validator_assigned_by_user_id,
        'validator_assigned_at': report.validator_assigned_at,
        'validated_by_user_id': report.validated_by_user_id,
        'validated_by_user': report.validated_by_user,
        'validated_at': report.validated_at,
        'validated_page_count': validated_page_count,
        'total_page_count': total_page_count,
        'can_finalize': can_finalize,
        'pages_validation': pages_validation,
    }


def get_report_validation_comment_snapshot(report: MonthlyReport) -> Dict[str, object]:
    """Return the latest validation comments and comment thread rows for a report."""
    comments = ReportValidationComment.objects.select_related('authored_by_user').filter(report=report).order_by(
        'page_key', '-updated_at', '-id'
    )
    latest_comments = {}
    threads = {}
    for comment in comments:
        thread_rows = threads.setdefault(comment.page_key, [])
        thread_rows.append({
            'page_key': comment.page_key,
            'comment_text': comment.comment_text,
            'authored_by_user_id': comment.authored_by_user_id,
            'authored_by_user_name': comment.authored_by_user.get_username() if comment.authored_by_user else None,
            'updated_at': comment.updated_at,
        })
        if comment.page_key not in latest_comments:
            latest_comments[comment.page_key] = comment.comment_text
    return {
        'validation_comments': latest_comments,
        'validation_comment_threads': threads,
    }


def upsert_report_validation_comments(*, report: MonthlyReport, comments_by_page: Dict[str, str], actor_user) -> None:
    """Persist validation comments without affecting business-content validation state."""
    if not comments_by_page:
        return

    allowed_page_keys = set(get_report_validation_page_keys(report))
    if allowed_page_keys:
        unknown_keys = sorted(set(comments_by_page).difference(allowed_page_keys))
        if unknown_keys:
            raise ValueError(f'Unknown report page key: {unknown_keys[0]}')

    now = dj_timezone.now()
    for page_key, text in comments_by_page.items():
        normalized_text = str(text or '')
        comment, _created = ReportValidationComment.objects.get_or_create(
            report=report,
            page_key=page_key,
            authored_by_user=actor_user,
            defaults={'comment_text': normalized_text},
        )
        if comment.comment_text != normalized_text:
            comment.comment_text = normalized_text
            comment.updated_at = now
            comment.save(update_fields=['comment_text', 'updated_at'])


def _resolve_validator_assigner_role(*, report: MonthlyReport, actor_user) -> Optional[str]:
    """Resolve effective assigner role for validator-selection and assignment."""
    role = _resolve_grantor_role(report=report, actor_user=actor_user)
    if role is not None:
        return role

    if not actor_user or not getattr(actor_user, 'is_authenticated', False):
        return None
    if not _is_user_in_report_scope(actor_user, report=report):
        return None

    from .models import has_user_role

    if Team.objects.filter(team_lead=actor_user).exists():
        return ReportWriteGrant.ROLE_TEAM_LEAD

    if (
        Team.objects.filter(manager=actor_user).exists()
        or has_user_role(actor_user, 'manager')
        or has_user_role(actor_user, 'admin')
    ):
        return ReportWriteGrant.ROLE_MANAGER

    return None


def _validation_superior_candidate_user_ids(report: MonthlyReport, actor_user) -> Set:
    """Return validator candidate user IDs for team-lead/manager assigners."""
    from .models import UserTeamAssignment

    candidate_ids: Set = set()

    actor_role = _resolve_validator_assigner_role(report=report, actor_user=actor_user)
    if actor_role == ReportWriteGrant.ROLE_TEAM_LEAD:
        actor_team_ids = set(_team_ids_led_by_user(actor_user))
    elif actor_role == ReportWriteGrant.ROLE_MANAGER:
        actor_team_ids = set(Team.objects.filter(manager=actor_user).values_list('id', flat=True))
    else:
        return candidate_ids

    if actor_team_ids:
        actor_scope_team_ids = set(actor_team_ids)
        for team in Team.objects.filter(id__in=actor_team_ids):
            actor_scope_team_ids.update(sub_team.id for sub_team in team.get_sub_teams())

        candidate_ids.update(
            UserTeamAssignment.objects.filter(team_id__in=actor_scope_team_ids).values_list('user_id', flat=True)
        )

    owner_team_ids = set(
        UserTeamAssignment.objects.filter(user_id=report.owner_user_id).values_list('team_id', flat=True)
    )
    supervisory_team_ids = set(owner_team_ids)
    for team in Team.objects.filter(id__in=owner_team_ids):
        supervisory_team_ids.update(parent_team.id for parent_team in team.get_parent_teams())

    if supervisory_team_ids:
        candidate_ids.update(
            Team.objects.filter(id__in=supervisory_team_ids)
            .exclude(team_lead_id=None)
            .values_list('team_lead_id', flat=True)
        )
        candidate_ids.update(
            Team.objects.filter(id__in=supervisory_team_ids)
            .exclude(manager_id=None)
            .values_list('manager_id', flat=True)
        )

    return candidate_ids


def _is_validation_eligible_user(report: MonthlyReport, user, actor_user=None) -> bool:
    """Return True when user is a valid validator for the report."""
    from .models import UserTeamAssignment

    if not user or not getattr(user, 'is_authenticated', False) or not getattr(user, 'is_active', False):
        return False
    if report.owner_user_id == getattr(user, 'id', None):
        return False
    if report.site is None:
        return False

    actor_role = _resolve_validator_assigner_role(report=report, actor_user=actor_user) if actor_user is not None else None
    if actor_role in {ReportWriteGrant.ROLE_TEAM_LEAD, ReportWriteGrant.ROLE_MANAGER}:
        return getattr(user, 'id', None) in _validation_superior_candidate_user_ids(report, actor_user)

    base_teams = []
    if report.site.team is not None:
        base_teams = [report.site.team]
    else:
        owner_team_ids = list(
            UserTeamAssignment.objects.filter(user_id=report.owner_user_id).values_list('team_id', flat=True)
        )
        if owner_team_ids:
            base_teams = list(Team.objects.filter(id__in=owner_team_ids))

    if not base_teams:
        return False

    base_team_ids = [team.id for team in base_teams]
    if UserTeamAssignment.objects.filter(user=user, team_id__in=base_team_ids).exists():
        return True

    supervisory_team_ids = set(base_team_ids)
    for team in base_teams:
        for parent_team in team.get_parent_teams():
            supervisory_team_ids.add(parent_team.id)

    return Team.objects.filter(id__in=supervisory_team_ids).filter(Q(team_lead=user) | Q(manager=user)).exists()


def assign_report_validator(*, report: MonthlyReport, validator_user, assigned_by_user) -> MonthlyReport:
    """Assign or reassign a validator for a report and reset validation state."""
    if not _is_validation_eligible_user(report, validator_user, actor_user=assigned_by_user):
        raise ValueError('Validator must be active, not the owner, and in the same team or supervisory chain')

    now = dj_timezone.now()
    reassignment = report.validator_user_id is not None and report.validator_user_id != getattr(validator_user, 'id', None)

    with transaction.atomic():
        report.validator_user = validator_user
        report.validator_assigned_by_user = assigned_by_user
        report.validator_assigned_at = now
        report.validation_status = MonthlyReport.VALIDATION_AWAITING
        if reassignment:
            report.validation_reopened_at = now
            report.validated_by_user = None
            report.validated_at = None
        report.save(
            update_fields=[
                'validator_user',
                'validator_assigned_by_user',
                'validator_assigned_at',
                'validation_status',
                'validation_reopened_at',
                'validated_by_user',
                'validated_at',
            ]
        )

        for row in ensure_report_validation_rows(report):
            if row.is_validated:
                row.is_validated = False
                row.validated_by_user = None
                row.validated_at = None
                row.reset_reason = 'validator_reassigned' if reassignment else 'validator_assigned'
                row.reset_at = now
                row.save(update_fields=['is_validated', 'validated_by_user', 'validated_at', 'reset_reason', 'reset_at'])

        ReportValidationEvent.objects.create(
            report=report,
            event_type=(
                ReportValidationEvent.EVENT_VALIDATOR_REASSIGNED
                if reassignment
                else ReportValidationEvent.EVENT_VALIDATOR_ASSIGNED
            ),
            event_by_user=assigned_by_user,
            metadata={
                'validator_user_id': getattr(validator_user, 'id', None),
                'reassignment': reassignment,
            },
        )

    return report


def _validation_page_rows(report: MonthlyReport) -> List[ReportPageValidationState]:
    """Return all validation rows for the report, ensuring canonical rows exist first."""
    return ensure_report_validation_rows(report)


def reset_report_page_validation_state(
    *,
    report: MonthlyReport,
    page_keys: Optional[List[str]] = None,
    reason: str = 'content_changed',
    actor_user=None,
) -> List[ReportPageValidationState]:
    """Reset validation state for one or more pages."""
    rows = _validation_page_rows(report)
    if page_keys:
        target_rows = [row for row in rows if row.page_key in set(page_keys)]
    else:
        target_rows = rows

    now = dj_timezone.now()
    for row in target_rows:
        if row.is_validated or row.validated_by_user_id or row.validated_at is not None:
            row.is_validated = False
            row.validated_by_user = None
            row.validated_at = None
            row.reset_reason = reason
            row.reset_at = now
            row.save(update_fields=['is_validated', 'validated_by_user', 'validated_at', 'reset_reason', 'reset_at'])
            ReportValidationEvent.objects.create(
                report=report,
                page_key=row.page_key,
                event_type=ReportValidationEvent.EVENT_PAGE_RESET,
                event_by_user=actor_user,
                metadata={'reason': reason},
            )

    if reason in {'content_changed', 'validator_reassigned', 'final_reopened'}:
        report.validation_status = MonthlyReport.VALIDATION_AWAITING
        report.validated_by_user = None
        report.validated_at = None
        if reason == 'final_reopened':
            report.validation_reopened_at = now
        report.save(update_fields=['validation_status', 'validated_by_user', 'validated_at', 'validation_reopened_at'])

    return target_rows


def mark_report_page_validation_state(
    *,
    report: MonthlyReport,
    page_key: str,
    is_validated: bool,
    actor_user,
    known_page_keys: Optional[List[str]] = None,
) -> ReportPageValidationState:
    """Mark a report page validated or unvalidated by the assigned validator."""
    if report.validator_user_id != getattr(actor_user, 'id', None):
        raise PermissionError('Only the assigned validator can mark report pages validated')

    page_keys = set(get_report_validation_page_keys(report))
    if page_keys and page_key not in page_keys:
        raise ValueError('Unknown report page key')

    if known_page_keys:
        ensure_report_validation_rows_for_keys(report, known_page_keys)

    row_map = {row.page_key: row for row in _validation_page_rows(report)}
    row = row_map.get(page_key)
    if row is None:
        row = ReportPageValidationState.objects.create(report=report, page_key=page_key)
    now = dj_timezone.now()

    if is_validated:
        row.is_validated = True
        row.validated_by_user = actor_user
        row.validated_at = now
        row.reset_reason = None
        row.reset_at = None
        row.save(update_fields=['is_validated', 'validated_by_user', 'validated_at', 'reset_reason', 'reset_at'])
        ReportValidationEvent.objects.create(
            report=report,
            page_key=page_key,
            event_type=ReportValidationEvent.EVENT_PAGE_VALIDATED,
            event_by_user=actor_user,
            metadata={'validated': True},
        )
    else:
        row.is_validated = False
        row.validated_by_user = None
        row.validated_at = None
        row.reset_reason = 'content_changed'
        row.reset_at = now
        row.save(update_fields=['is_validated', 'validated_by_user', 'validated_at', 'reset_reason', 'reset_at'])
        ReportValidationEvent.objects.create(
            report=report,
            page_key=page_key,
            event_type=ReportValidationEvent.EVENT_PAGE_RESET,
            event_by_user=actor_user,
            metadata={'reason': 'manual_unvalidate'},
        )

    rows = _validation_page_rows(report)
    if rows and all(row.is_validated for row in rows):
        report.validation_status = MonthlyReport.VALIDATION_VALIDATED
        report.validated_by_user = actor_user
        report.validated_at = now
        report.save(update_fields=['validation_status', 'validated_by_user', 'validated_at'])
        ReportValidationEvent.objects.create(
            report=report,
            event_type=ReportValidationEvent.EVENT_REPORT_VALIDATED,
            event_by_user=actor_user,
            metadata={'validated_page_count': len(rows), 'total_page_count': len(rows)},
        )
    else:
        report.validation_status = MonthlyReport.VALIDATION_AWAITING
        report.validated_by_user = None
        report.validated_at = None
        report.save(update_fields=['validation_status', 'validated_by_user', 'validated_at'])

    return row


def can_user_assign_report_validator(report: MonthlyReport, user) -> bool:
    """Return True when the user may assign a validator for this report."""
    return _resolve_validator_assigner_role(report=report, actor_user=user) is not None


def get_report_validation_candidate_users(report: MonthlyReport, actor_user=None) -> List[Dict[str, object]]:
    """Return active users eligible to be assigned as report validators."""
    from django.contrib.auth import get_user_model

    user_model = get_user_model()
    users_qs = user_model.objects.filter(is_active=True).exclude(id=report.owner_user_id)

    actor_role = _resolve_validator_assigner_role(report=report, actor_user=actor_user) if actor_user is not None else None

    if actor_role in {ReportWriteGrant.ROLE_TEAM_LEAD, ReportWriteGrant.ROLE_MANAGER}:
        candidate_ids = _validation_superior_candidate_user_ids(report, actor_user)
        users_qs = users_qs.filter(id__in=candidate_ids)

    users = users_qs.order_by('username')

    candidates = []
    for user in users:
        if _is_validation_eligible_user(report, user, actor_user=actor_user):
            candidates.append({
                'id': str(user.id),
                'username': user.get_username(),
            })
    return candidates


def _next_report_version_number(report: MonthlyReport) -> int:
    latest = report.versions.order_by('-version_number').first()
    return 1 if latest is None else latest.version_number + 1


def create_report_version(
    report: MonthlyReport,
    version_kind: str,
    comments: Optional[Dict[str, str]] = None,
    derived_from_version: Optional[MonthlyReportVersion] = None,
    actor_user=None,
) -> MonthlyReportVersion:
    """Create an immutable report version and update report pointers."""
    version = MonthlyReportVersion.objects.create(
        report=report,
        version_number=_next_report_version_number(report),
        version_kind=version_kind,
        derived_from_version=derived_from_version,
    )

    for visual_key, text in (comments or {}).items():
        ReportComment.objects.create(
            report_version=version,
            visual_key=str(visual_key),
            text=str(text or ''),
        )

    report.current_version = version
    if version_kind in {MonthlyReportVersion.KIND_FINAL, MonthlyReportVersion.KIND_REPLACEMENT_FINAL}:
        report.current_status = MonthlyReport.STATUS_FINAL
        report.current_final_version = version
    else:
        report.current_status = MonthlyReport.STATUS_DRAFT

    if actor_user is not None:
        if report.owner_user_id is None:
            report.owner_user = actor_user
        if report.created_by_user_id is None:
            report.created_by_user = actor_user
        report.last_modified_by_user = actor_user
    report.last_modified_at = dj_timezone.now()

    report.save(
        update_fields=[
            'current_version',
            'current_final_version',
            'current_status',
            'owner_user',
            'created_by_user',
            'last_modified_by_user',
            'last_modified_at',
            'updated_at',
        ]
    )
    return version


def get_report_write_grant(report: MonthlyReport, user) -> Optional[ReportWriteGrant]:
    """Return active write grant for the report and user if present."""
    if not user or not getattr(user, 'is_authenticated', False):
        return None

    grant = ReportWriteGrant.objects.filter(report=report, granted_user=user, is_active=True).first()
    if grant is None:
        return None

    # Submit-time eligibility re-check: inactive or out-of-scope delegates lose effective access.
    if not _is_grant_delegate_still_eligible(report=report, grant=grant):
        return None
    return grant


def _team_scope_ids_for_site(site: Site) -> Set:
    if site.team_id is None:
        return set()

    team_ids = {site.team_id}
    for sub_team in site.team.get_sub_teams():
        team_ids.add(sub_team.id)
    parent_teams = site.team.get_parent_teams()
    for parent_team in parent_teams:
        team_ids.add(parent_team.id)

    # Organisation scope should include sibling branches under each ancestor.
    for parent_team in parent_teams:
        for sibling_branch_team in parent_team.get_sub_teams():
            team_ids.add(sibling_branch_team.id)

    return team_ids


def _scope_team_ids_for_report(report: MonthlyReport) -> Set:
    """Return delegation scope team IDs for a report.

    Primary scope is report.site.team hierarchy. If site team is missing,
    fall back to the report owner's assigned teams and their hierarchy.
    """
    if report.site and report.site.team_id:
        return _team_scope_ids_for_site(report.site)

    from .models import UserTeamAssignment

    owner = report.owner_user
    if owner is None:
        return set()

    scope_team_ids: Set = set()
    owner_team_ids = UserTeamAssignment.objects.filter(user=owner).values_list('team_id', flat=True)
    for team in Team.objects.filter(id__in=owner_team_ids):
        scope_team_ids.add(team.id)
        for sub_team in team.get_sub_teams():
            scope_team_ids.add(sub_team.id)
        for parent_team in team.get_parent_teams():
            scope_team_ids.add(parent_team.id)

    return scope_team_ids


def _team_ids_led_by_user(user) -> Set:
    """Return team IDs where the user is explicitly configured as team lead."""
    if not user or not getattr(user, 'is_authenticated', False):
        return set()
    return set(Team.objects.filter(team_lead=user).values_list('id', flat=True))


def _is_team_lead_for_report(user, *, report: MonthlyReport) -> bool:
    """Return True when user is a lead for the report owner's/site team context."""
    from .models import UserTeamAssignment

    lead_team_ids = _team_ids_led_by_user(user)
    if not lead_team_ids:
        return False

    if report.site and report.site.team_id and report.site.team_id in lead_team_ids:
        return True

    owner = report.owner_user
    if owner is None:
        return False

    owner_team_ids = set(UserTeamAssignment.objects.filter(user=owner).values_list('team_id', flat=True))
    return bool(owner_team_ids.intersection(lead_team_ids))


def _is_user_assigned_to_scope_strict(user, *, site: Site) -> bool:
    """Scope membership check without implicit admin bypass for delegation decisions."""
    from .models import UserTeamAssignment

    if not user or not getattr(user, 'is_authenticated', False):
        return False
    if site.team_id is None:
        return False

    scope_team_ids = _team_scope_ids_for_site(site)

    # Directly assigned on the report team itself.
    if user.id in {site.team.team_lead_id, site.team.manager_id}:
        return True

    # Leads/managers attached to parent/sub-team nodes in the same hierarchy are in scope.
    if Team.objects.filter(id__in=scope_team_ids).filter(Q(team_lead=user) | Q(manager=user)).exists():
        return True
    return UserTeamAssignment.objects.filter(user=user, team_id__in=scope_team_ids).exists()


def _is_user_in_report_scope(user, *, report: MonthlyReport) -> bool:
    """Return True when user belongs to report delegation scope."""
    from .models import UserTeamAssignment

    if not user or not getattr(user, 'is_authenticated', False):
        return False

    scope_team_ids = _scope_team_ids_for_report(report)
    if not scope_team_ids:
        return False

    if Team.objects.filter(id__in=scope_team_ids).filter(Q(team_lead=user) | Q(manager=user)).exists():
        return True

    return UserTeamAssignment.objects.filter(user=user, team_id__in=scope_team_ids).exists()


def _users_share_any_team(user_a, user_b) -> bool:
    from .models import UserTeamAssignment

    if not user_a or not user_b:
        return False

    team_ids_a = set(UserTeamAssignment.objects.filter(user=user_a).values_list('team_id', flat=True))
    if not team_ids_a:
        return False
    return UserTeamAssignment.objects.filter(user=user_b, team_id__in=team_ids_a).exists()


def _resolve_grantor_role(*, report: MonthlyReport, actor_user) -> Optional[str]:
    if not actor_user or not getattr(actor_user, 'is_authenticated', False):
        return None
    if report.owner_user_id == actor_user.id:
        return ReportWriteGrant.ROLE_OWNER

    from .models import has_user_role

    # Admins can manage delegation for reports and are tracked under manager-equivalent authority.
    if has_user_role(actor_user, 'admin'):
        return ReportWriteGrant.ROLE_MANAGER

    scope_team_ids = _scope_team_ids_for_report(report)
    is_scope_team_lead = _is_team_lead_for_report(actor_user, report=report)
    is_scope_manager = Team.objects.filter(id__in=scope_team_ids, manager=actor_user).exists()

    if _is_user_in_report_scope(actor_user, report=report):
        if (
            actor_user.id == getattr(report.site.team, 'team_lead_id', None)
            or is_scope_team_lead
        ):
            return ReportWriteGrant.ROLE_TEAM_LEAD
        if (
            actor_user.id == getattr(report.site.team, 'manager_id', None)
            or is_scope_manager
            or has_user_role(actor_user, 'manager')
        ):
            return ReportWriteGrant.ROLE_MANAGER

    return None


def _validate_delegate_scope_for_grant(*, report: MonthlyReport, granted_user, granted_by_role: str, granted_by_user=None):
    from .models import has_user_role

    if not granted_user.is_active:
        raise ValueError('Cannot grant write access to an inactive user')

    if granted_by_user is not None and has_user_role(granted_by_user, 'admin'):
        return

    if granted_by_role == ReportWriteGrant.ROLE_OWNER:
        owner = report.owner_user
        if owner is None or not _users_share_any_team(owner, granted_user):
            raise ValueError('Owner grants require delegate to be in the same team')
        return

    if granted_by_role == ReportWriteGrant.ROLE_TEAM_LEAD:
        if granted_by_user is not None and getattr(granted_user, 'id', None) == getattr(granted_by_user, 'id', None):
            return

        lead_team_ids = _team_ids_led_by_user(granted_by_user)
        if not lead_team_ids:
            raise ValueError('Team lead grants require team-lead assignment to a team')

        from .models import UserTeamAssignment

        if not UserTeamAssignment.objects.filter(user=granted_user, team_id__in=lead_team_ids).exists():
            raise ValueError("Team lead grants require delegate to be in the team lead's team")
        return

    if not _is_user_in_report_scope(granted_user, report=report):
        raise ValueError('Delegate must be in the same organisation scope as the report')


def _is_grant_delegate_still_eligible(*, report: MonthlyReport, grant: ReportWriteGrant) -> bool:
    user = grant.granted_user
    if not user or not user.is_active:
        return False

    if grant.granted_by_role == ReportWriteGrant.ROLE_OWNER:
        owner = report.owner_user
        return owner is not None and _users_share_any_team(owner, user)

    return _is_user_in_report_scope(user, report=report)


def _can_revoke_report_grant(*, report: MonthlyReport, grant: ReportWriteGrant, revoked_by) -> bool:
    if report.owner_user_id == getattr(revoked_by, 'id', None):
        return True
    if grant.granted_by_id == getattr(revoked_by, 'id', None):
        return True
    return _resolve_grantor_role(report=report, actor_user=revoked_by) in {
        ReportWriteGrant.ROLE_TEAM_LEAD,
        ReportWriteGrant.ROLE_MANAGER,
    }


def _record_report_delegation_event(
    *,
    report: MonthlyReport,
    delegate_user,
    action: str,
    action_by_user,
    action_by_role: str,
    resolution_basis: Optional[str] = None,
    correlation_key=None,
    notes: str = '',
) -> ReportWriteDelegationEvent:
    return ReportWriteDelegationEvent.objects.create(
        report=report,
        delegate_user=delegate_user,
        action=action,
        action_by_user=action_by_user,
        action_by_role=action_by_role,
        resolution_basis=resolution_basis,
        correlation_key=correlation_key,
        notes=notes or '',
    )


def get_report_delegation_history(report: MonthlyReport):
    return ReportWriteDelegationEvent.objects.filter(report=report).select_related(
        'delegate_user',
        'action_by_user',
    )


def get_active_report_write_grant_for_update(*, report: MonthlyReport, granted_user) -> Optional[ReportWriteGrant]:
    """Select active grant row with lock for transaction-safe grant/revoke decisions."""
    return (
        ReportWriteGrant.objects.select_for_update()
        .filter(report=report, granted_user=granted_user, is_active=True)
        .first()
    )


def get_report_delegation_visibility_rows(report: MonthlyReport) -> List[Dict[str, object]]:
    """Return active delegation rows for report visibility responses."""
    rows = []
    for grant in ReportWriteGrant.objects.filter(report=report, is_active=True).select_related(
        'granted_user',
        'granted_by',
    ):
        rows.append(
            {
                'delegate_user_id': str(grant.granted_user_id),
                'delegate_user': grant.granted_user.get_username(),
                'granted_by_user_id': str(grant.granted_by_id) if grant.granted_by_id else '',
                'granted_by_user': grant.granted_by.get_username() if grant.granted_by else '',
                'granted_by_role': grant.granted_by_role,
                'granted_at': grant.granted_at,
                'is_active': grant.is_active,
            }
        )
    return rows


def can_user_manage_report_delegations(report: MonthlyReport, user) -> bool:
    """Return True when user can grant/revoke delegation for a report."""
    return _resolve_grantor_role(report=report, actor_user=user) is not None


def get_user_report_delegation_role(report: MonthlyReport, user) -> Optional[str]:
    """Return effective delegation role for user on report, if any."""
    return _resolve_grantor_role(report=report, actor_user=user)


def get_report_delegation_role_hint(report: MonthlyReport, user) -> str:
    """Return UI hint describing delegation scope for the current user role."""
    role = get_user_report_delegation_role(report, user)
    if role == ReportWriteGrant.ROLE_OWNER:
        return 'You can grant as owner to active users in your same team.'
    if role == ReportWriteGrant.ROLE_TEAM_LEAD:
        return 'You can grant as team lead to active users in your own team, including yourself.'
    if role == ReportWriteGrant.ROLE_MANAGER:
        return 'You can grant as manager to active users in your organisation, including yourself.'
    return ''


def get_report_delegation_candidate_users(report: MonthlyReport, actor_user) -> List[Dict[str, object]]:
    """Return active user options that the actor can delegate to for this report."""
    from .models import UserTeamAssignment
    from django.contrib.auth import get_user_model

    from .models import has_user_role

    role = _resolve_grantor_role(report=report, actor_user=actor_user)
    if role is None:
        return []

    include_owner_for_final_regrant = (
        report.current_status == MonthlyReport.STATUS_FINAL
        and report.validation_status == MonthlyReport.VALIDATION_VALIDATED
        and role in {ReportWriteGrant.ROLE_TEAM_LEAD, ReportWriteGrant.ROLE_MANAGER}
    )

    user_model = get_user_model()
    base_qs = user_model.objects.filter(is_active=True)

    if has_user_role(actor_user, 'admin'):
        users = base_qs.distinct().order_by('username')
        if not include_owner_for_final_regrant:
            users = users.exclude(id=report.owner_user_id)
        return [{'id': str(user.id), 'username': user.get_username()} for user in users]

    if role == ReportWriteGrant.ROLE_OWNER:
        owner_team_ids = UserTeamAssignment.objects.filter(user=report.owner_user).values_list('team_id', flat=True)
        base_qs = base_qs.filter(team_assignments__team_id__in=owner_team_ids)
    elif role == ReportWriteGrant.ROLE_TEAM_LEAD:
        lead_team_ids = _team_ids_led_by_user(actor_user)
        scoped_user_ids = set(UserTeamAssignment.objects.filter(team_id__in=lead_team_ids).values_list('user_id', flat=True))
        scoped_user_ids.add(actor_user.id)
        base_qs = base_qs.filter(id__in=scoped_user_ids)
    else:
        scope_team_ids = _scope_team_ids_for_report(report)
        scoped_user_ids = set(UserTeamAssignment.objects.filter(team_id__in=scope_team_ids).values_list('user_id', flat=True))
        scoped_user_ids.add(actor_user.id)
        base_qs = base_qs.filter(id__in=scoped_user_ids)

    users = base_qs.distinct().order_by('username')
    if not include_owner_for_final_regrant:
        users = users.exclude(id=report.owner_user_id)

    return [{'id': str(user.id), 'username': user.get_username()} for user in users]


def get_report_access_mode(report: MonthlyReport, user) -> str:
    """Return one of owner, collaborator, validator, admin, or read_only for report access."""
    if not user or not getattr(user, 'is_authenticated', False):
        return 'read_only'

    # Finalized validated reports are locked until a superior re-grants write access.
    if (
        report.current_status == MonthlyReport.STATUS_FINAL
        and report.validation_status == MonthlyReport.VALIDATION_VALIDATED
    ):
        final_write_grant = get_report_write_grant(report, user)
        if (
            final_write_grant is None
            or final_write_grant.granted_by_role not in {ReportWriteGrant.ROLE_TEAM_LEAD, ReportWriteGrant.ROLE_MANAGER}
        ):
            return 'read_only'

    if report.owner_user_id == user.id:
        return 'owner'
    if report.validator_user_id == user.id:
        return 'validator'
    if get_report_write_grant(report, user) is not None:
        return 'collaborator'
    if user.is_staff or user.is_superuser:
        return 'admin'
    return 'read_only'


def user_can_write_report(report: MonthlyReport, user) -> bool:
    """Return True if user is authorized to edit report content."""
    return get_report_access_mode(report, user) in {'owner', 'collaborator', 'validator', 'admin'}


def grant_report_write_access(*, report: MonthlyReport, granted_user, granted_by) -> ReportWriteGrant:
    """Grant report write access with owner/team-lead/manager delegation rules."""
    granted_by_role = _resolve_grantor_role(report=report, actor_user=granted_by)
    if granted_by_role is None:
        raise PermissionError('Only report owner, same-organisation team lead, or manager can grant write access')

    if granted_by_role == ReportWriteGrant.ROLE_OWNER and getattr(granted_user, 'id', None) == report.owner_user_id:
        raise ValueError('Owner already has write access')

    _validate_delegate_scope_for_grant(
        report=report,
        granted_user=granted_user,
        granted_by_role=granted_by_role,
        granted_by_user=granted_by,
    )

    with transaction.atomic():
        existing_active = get_active_report_write_grant_for_update(report=report, granted_user=granted_user)
        if existing_active:
            raise ValueError('Write access already granted to this user')

        existing = (
            ReportWriteGrant.objects.select_for_update()
            .filter(report=report, granted_user=granted_user)
            .order_by('-granted_at')
            .first()
        )
        if existing is None:
            grant = ReportWriteGrant.objects.create(
                report=report,
                granted_user=granted_user,
                granted_by=granted_by,
                granted_by_role=granted_by_role,
                is_active=True,
            )
        else:
            grant = existing
            grant.granted_by = granted_by
            grant.granted_by_role = granted_by_role
            grant.granted_at = dj_timezone.now()
            grant.revoked_by = None
            grant.revoked_by_role = None
            grant.revoked_at = None
            grant.is_active = True
            grant.save(
                update_fields=[
                    'granted_by',
                    'granted_by_role',
                    'granted_at',
                    'revoked_by',
                    'revoked_by_role',
                    'revoked_at',
                    'is_active',
                ]
            )

        _record_report_delegation_event(
            report=report,
            delegate_user=granted_user,
            action=ReportWriteDelegationEvent.ACTION_GRANT,
            action_by_user=granted_by,
            action_by_role=granted_by_role,
        )
        return grant


def revoke_report_write_access(*, report: MonthlyReport, granted_user, revoked_by) -> Optional[ReportWriteGrant]:
    """Revoke active write access with owner/original-grantor/lead-manager authority."""
    revoked_by_role = _resolve_grantor_role(report=report, actor_user=revoked_by)

    with transaction.atomic():
        grant = get_active_report_write_grant_for_update(report=report, granted_user=granted_user)
        if not grant:
            return None

        if not _can_revoke_report_grant(report=report, grant=grant, revoked_by=revoked_by):
            raise PermissionError('Only report owner, original grantor, or same-organisation team lead/manager can revoke write access')

        grant.is_active = False
        grant.revoked_by = revoked_by
        grant.revoked_by_role = revoked_by_role or ReportWriteGrant.ROLE_OWNER
        grant.revoked_at = dj_timezone.now()
        grant.save(update_fields=['is_active', 'revoked_by', 'revoked_by_role', 'revoked_at'])

        _record_report_delegation_event(
            report=report,
            delegate_user=granted_user,
            action=ReportWriteDelegationEvent.ACTION_REVOKE,
            action_by_user=revoked_by,
            action_by_role=grant.revoked_by_role,
            resolution_basis=ReportWriteDelegationEvent.RESOLUTION_LAST_WRITE_WINS,
        )

        return grant


def _is_user_assigned_to_scope(user, *, site: Site) -> bool:
    """Return True when user is in the same site/team scope or has platform admin status."""
    if user.is_staff or user.is_superuser:
        return True
    from .models import UserTeamAssignment

    if site.team_id is None:
        return False

    if user.id in {site.team.team_lead_id, site.team.manager_id}:
        return True

    team_ids = {site.team_id}
    for sub_team in site.team.get_sub_teams():
        team_ids.add(sub_team.id)
    for parent_team in site.team.get_parent_teams():
        team_ids.add(parent_team.id)

    return UserTeamAssignment.objects.filter(user=user, team_id__in=team_ids).exists()


def _fallback_candidates_for_report(report: MonthlyReport) -> List:
    """Build ordered fallback candidates: team lead, manager, then scoped admin."""
    site = report.site
    if site.team_id is None:
        return []

    candidates = []
    if site.team_id:
        if site.team.team_lead and site.team.team_lead.is_active:
            candidates.append(site.team.team_lead)
        if site.team.manager and site.team.manager.is_active and site.team.manager_id != getattr(site.team.team_lead, 'id', None):
            candidates.append(site.team.manager)

    from .models import RoleAssignment, UserTeamAssignment
    scoped_admin_ids = set()
    if site.team_id:
        team_ids = {site.team_id}
        team_ids.update([team.id for team in site.team.get_sub_teams()])
        team_ids.update([team.id for team in site.team.get_parent_teams()])
        scoped_admin_ids.update(
            UserTeamAssignment.objects.filter(team_id__in=team_ids).values_list('user_id', flat=True)
        )

    from django.contrib.auth import get_user_model
    UserModel = get_user_model()

    role_admin_ids = set(
        RoleAssignment.objects.filter(role_name='admin').values_list('user_id', flat=True)
    )
    if scoped_admin_ids:
        role_admin_ids = role_admin_ids.intersection(scoped_admin_ids)
    for user in UserModel.objects.filter(id__in=role_admin_ids, is_active=True).order_by('id'):
        if user.id not in {candidate.id for candidate in candidates}:
            candidates.append(user)

    return candidates


def transfer_report_ownership(
    *,
    report: MonthlyReport,
    new_owner,
    transfer_mode: str,
    transfer_reason: str = '',
    approval_record: Optional[ReportOwnershipUnavailabilityApproval] = None,
    executed_by=None,
) -> ReportOwnershipTransferEvent:
    """Transfer report ownership and persist transfer event."""
    previous_owner = report.owner_user
    if previous_owner and previous_owner.id == new_owner.id:
        raise ValueError('new owner must be different from current owner')

    report.owner_user = new_owner
    report.last_modified_by_user = executed_by or new_owner
    report.last_modified_at = dj_timezone.now()
    report.save(update_fields=['owner_user', 'last_modified_by_user', 'last_modified_at', 'updated_at'])

    if previous_owner:
        ReportWriteGrant.objects.get_or_create(
            report=report,
            granted_user=previous_owner,
            is_active=True,
            defaults={'granted_by': new_owner},
        )

    return ReportOwnershipTransferEvent.objects.create(
        report=report,
        from_owner=previous_owner,
        to_owner=new_owner,
        transfer_mode=transfer_mode,
        transfer_reason=transfer_reason or '',
        approval_record=approval_record,
        executed_by=executed_by,
    )


def approve_owner_unavailability_and_transfer(*, report: MonthlyReport, owner_user, approved_by, reason: str):
    """Approve owner unavailability and execute deterministic fallback transfer."""
    from .models import has_user_role

    if not reason.strip():
        raise ValueError('approval reason is required')
    if report.owner_user_id != owner_user.id:
        raise ValueError('owner_user must match report owner')
    if not has_user_role(approved_by, 'team_lead') and approved_by.id != getattr(report.site.team, 'team_lead_id', None):
        raise PermissionError('team lead approval is required')
    if report.site.team_id is None:
        raise LookupError('Fallback transfer requires site team scope to be configured')

    approval = ReportOwnershipUnavailabilityApproval.objects.create(
        report=report,
        owner_user=owner_user,
        approved_by=approved_by,
        approval_reason=reason.strip(),
        status=ReportOwnershipUnavailabilityApproval.STATUS_APPROVED,
    )

    for candidate in _fallback_candidates_for_report(report):
        if candidate.id == owner_user.id:
            continue
        if not candidate.is_active:
            continue
        if not _is_user_assigned_to_scope(candidate, site=report.site):
            continue

        transfer_event = transfer_report_ownership(
            report=report,
            new_owner=candidate,
            transfer_mode=ReportOwnershipTransferEvent.MODE_AUTO_FALLBACK,
            transfer_reason=f'Fallback transfer after team-lead approval: {reason.strip()}',
            approval_record=approval,
            executed_by=approved_by,
        )
        return approval, transfer_event

    raise LookupError('No eligible fallback ownership candidate found in required order')


def get_previous_month_final_version(site: Site, reporting_month: str) -> Optional[MonthlyReportVersion]:
    """Return the prior month final version for the same site, if available."""
    start, _ = reporting_month_bounds(reporting_month)
    previous_month = shift_months(start, -1).strftime('%Y-%m')
    previous_report = MonthlyReport.objects.filter(site=site, reporting_month=previous_month).first()
    if not previous_report:
        return None
    return previous_report.current_final_version


def carry_forward_comments_from_previous_final(
    report: MonthlyReport,
    report_version: MonthlyReportVersion,
) -> int:
    """Copy previous-month final comments into this version as reference comments.

    A comment is (re)tagged as a reference copy unless the version already has
    a saved comment for that visual key with different text - which means the
    user edited the prefilled reference comment before saving, so their edit
    is preserved as-is.
    """
    previous_final = get_previous_month_final_version(report.site, report.reporting_month)
    if previous_final is None:
        return 0

    existing_comments = {comment.visual_key: comment for comment in report_version.comments.all()}

    copied = 0
    for previous_comment in previous_final.comments.all().order_by('visual_key'):
        existing = existing_comments.get(previous_comment.visual_key)
        if existing is not None and existing.text != previous_comment.text:
            continue

        ReportComment.objects.update_or_create(
            report_version=report_version,
            visual_key=previous_comment.visual_key,
            defaults={
                'text': previous_comment.text,
                'is_reference_copy': True,
                'source_reporting_month': previous_final.report.reporting_month,
                'source_version': previous_final,
            },
        )
        copied += 1

    return copied


def set_import_run_status(import_run: ImportRun, status: str, error_details: Optional[Dict] = None) -> None:
    """Persist import run status transitions and completion timestamps."""
    import_run.status = status
    if status in {ImportRun.STATUS_SUCCESS, ImportRun.STATUS_PARTIAL_FAILURE, ImportRun.STATUS_FAILED}:
        import_run.completed_at = dj_timezone.now()
    if error_details is not None:
        import_run.error_details = error_details
    import_run.save(
        update_fields=['status', 'completed_at', 'error_details', 'updated_at']
        if error_details is not None
        else ['status', 'completed_at', 'updated_at']
    )


def upsert_halfhourly_record(import_run: ImportRun, supply: Supply, row: Dict) -> Tuple[HalfHourlyConsumption, bool]:
    start = parse_utc_datetime(row.get('startDate') or row.get('periodStart') or row.get('start_date'))
    end = parse_utc_datetime(row.get('endDate') or row.get('periodEnd') or row.get('end_date'))
    breakdown = row.get('combinedBreakdown') or {}
    hh_value = None
    if isinstance(breakdown, dict):
        hh_value = breakdown.get('hh')
    if hh_value is None:
        hh_value = row.get('consumption', 0)
    value = Decimal(str(hh_value or 0))
    return HalfHourlyConsumption.objects.update_or_create(
        supply=supply,
        source_period_start=start,
        source_period_end=end,
        defaults={
            'import_run': import_run,
            'canonical_month_key': canonical_month_key(end),
            'consumption': value,
        },
    )


def upsert_monthly_record(import_run: ImportRun, supply: Supply, row: Dict) -> Tuple[MonthlyConsumption, bool]:
    start = parse_utc_datetime(row.get('startDate') or row.get('periodStart') or row.get('start_date'))
    end = parse_utc_datetime(row.get('endDate') or row.get('periodEnd') or row.get('end_date'))
    value = Decimal(str(row.get('consumption', 0) or 0))
    breakdown = row.get('combinedBreakdown') or row.get('breakdown') or {}
    sources = row.get('sources') or []
    return MonthlyConsumption.objects.update_or_create(
        supply=supply,
        source_period_start=start,
        source_period_end=end,
        defaults={
            'import_run': import_run,
            'canonical_month_key': canonical_month_key(end),
            'consumption': value,
            'breakdown': breakdown if isinstance(breakdown, dict) else {},
            'sources': sources if isinstance(sources, list) else [],
        },
    )


def upsert_invoice_record(import_run: ImportRun, supply: Supply, row: Dict) -> Tuple[InvoiceCost, bool]:
    values = row.get('values') if isinstance(row.get('values'), dict) else {}

    def pick(*keys: str):
        for key in keys:
            if values.get(key) is not None:
                return values.get(key)
            if row.get(key) is not None:
                return row.get(key)
        return None

    start_raw = pick(
        'startDate',
        'periodStart',
        'fromDate',
        'start_date',
        'period_start',
        'invoiceDate',
        'date',
    )
    end_raw = pick(
        'endDate',
        'periodEnd',
        'toDate',
        'end_date',
        'period_end',
    ) or start_raw
    start = parse_utc_datetime(start_raw)
    end = parse_utc_datetime(end_raw)
    cost = Decimal(str(pick('netTotalCost', 'netCost', 'cost', 'amount') or 0))
    metadata = {
        'invoiceDate': pick('invoiceDate', 'date'),
        'invoiceNumber': pick('invoiceNumber', 'number'),
        'status': pick('status'),
        'financialStatus': pick('financialStatus'),
        'completed': pick('completed'),
        'type': pick('type'),
    }
    return InvoiceCost.objects.update_or_create(
        supply=supply,
        source_period_start=start,
        source_period_end=end,
        defaults={
            'import_run': import_run,
            'canonical_month_key': canonical_month_key(end),
            'cost': cost,
            'invoice_metadata': {k: v for k, v in metadata.items() if v is not None},
        },
    )


class ConsumptionImportService:
    """Orchestrates usage/invoice imports for selected supplies and reporting month."""

    def __init__(self, api_client: Optional[EtainablApiClient] = None):
        self.client = api_client or EtainablApiClient()
        self.retry_count = max(0, int(settings.CONSUMPTION_IMPORT_RETRY_COUNT))
        self.retry_backoff = max(1, int(settings.CONSUMPTION_IMPORT_RETRY_BACKOFF_SECONDS))
        app_settings = SettingsConfigService.get_settings()
        self.invoice_page_limit = max(1, int(app_settings.invoice_page_limit or 100))
        self.invoice_start_page = max(1, int(app_settings.invoice_start_page or 1))

    def run(self, supply_external_ids: List[str], reporting_month: str, refresh_mode: bool = True) -> ImportRun:
        import_run = ImportRun.objects.create(
            selected_supply_ids=supply_external_ids,
            reporting_month=reporting_month,
            status=ImportRun.STATUS_IN_PROGRESS,
            affected_supply_count=len(supply_external_ids),
        )

        records_imported = 0
        records_failed = 0
        retry_used = 0
        errors: Dict[str, str] = {}
        outcomes: List[Dict] = []

        supplies = list(Supply.objects.filter(external_id__in=supply_external_ids).order_by('id'))
        supply_by_external_id = {s.external_id: s for s in supplies}

        for supply_external_id in supply_external_ids:
            supply = supply_by_external_id.get(supply_external_id)
            if not supply:
                records_failed += 1
                errors[supply_external_id] = 'Supply not found'
                outcomes.append({
                    'supply_id': supply_external_id,
                    'status': 'failed',
                    'failure_reason': 'Supply not found',
                })
                continue

            try:
                outcome, imported_count, retries_for_supply, failed_ops = self._import_supply(
                    import_run,
                    supply,
                    reporting_month,
                    refresh_mode=refresh_mode,
                )
                records_imported += imported_count
                retry_used += retries_for_supply
                records_failed += failed_ops
                outcomes.extend(outcome)
            except Exception as exc:  # pylint: disable=broad-except
                records_failed += 1
                errors[supply_external_id] = str(exc)
                outcomes.append({
                    'supply_id': supply.external_id,
                    'status': 'failed',
                    'failure_reason': str(exc),
                })

        import_run.records_imported = records_imported
        import_run.records_failed = records_failed
        import_run.retry_count = retry_used
        import_run.error_details = errors
        import_run.outcome_details = outcomes

        if records_failed == 0:
            import_run.status = ImportRun.STATUS_SUCCESS
        elif records_imported > 0:
            import_run.status = ImportRun.STATUS_PARTIAL_FAILURE
        else:
            import_run.status = ImportRun.STATUS_FAILED

        import_run.completed_at = dj_timezone.now()
        import_run.save()
        return import_run

    def _import_supply(
        self,
        import_run: ImportRun,
        supply: Supply,
        reporting_month: str,
        refresh_mode: bool,
    ) -> Tuple[List[Dict], int, int, int]:
        outcomes = []
        imported_count = 0
        retries_used = 0
        failed_ops = 0

        hh_windows = get_halfhourly_windows(reporting_month)
        for start, end in hh_windows:
            if not refresh_mode and HalfHourlyConsumption.objects.filter(
                supply=supply,
                source_period_start__gte=start,
                source_period_start__lt=end,
            ).exists():
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'halfhourly',
                    'request_window': {'start': start.isoformat(), 'end': end.isoformat()},
                    'attempt_count': 0,
                    'retry_used': False,
                    'response_code': None,
                    'status': 'skipped',
                    'reason': 'cached_data_reused',
                })
                continue
            try:
                rows, retries, account_id_used = self._fetch_consumption_for_supply(
                    supply=supply,
                    start_date=format_api_datetime(start),
                    end_date=format_api_datetime(end),
                    granularity='halfhourly',
                    source='combined',
                )
                retries_used += retries
                payload_rows = rows.get('data') if isinstance(rows, dict) else []
                payload_rows = payload_rows if isinstance(payload_rows, list) else []
                with transaction.atomic():
                    for row in payload_rows:
                        upsert_halfhourly_record(import_run, supply, row)
                        imported_count += 1
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'halfhourly',
                    'request_window': {'start': start.isoformat(), 'end': end.isoformat()},
                    'attempt_count': retries + 1,
                    'retry_used': retries > 0,
                    'response_code': 200,
                    'account_id': account_id_used,
                    'status': 'success',
                })
            except Exception as exc:  # pylint: disable=broad-except
                failed_ops += 1
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'halfhourly',
                    'request_window': {'start': start.isoformat(), 'end': end.isoformat()},
                    'attempt_count': self.retry_count + 1,
                    'retry_used': self.retry_count > 0,
                    'response_code': None,
                    'status': 'failed',
                    'failure_reason': str(exc),
                })

        monthly_start, monthly_end = get_monthly_window(reporting_month)
        if not refresh_mode and MonthlyConsumption.objects.filter(
            supply=supply,
            source_period_start__gte=monthly_start,
            source_period_start__lt=monthly_end,
        ).exists():
            outcomes.append({
                'supply_id': supply.external_id,
                'data_type': 'monthly',
                'request_window': {'start': monthly_start.isoformat(), 'end': monthly_end.isoformat()},
                'attempt_count': 0,
                'retry_used': False,
                'response_code': None,
                'status': 'skipped',
                'reason': 'cached_data_reused',
            })
        else:
            try:
                monthly_payload, retries, account_id_used = self._fetch_consumption_for_supply(
                    supply=supply,
                    start_date=format_api_datetime(monthly_start),
                    end_date=format_api_datetime(monthly_end),
                    granularity='monthly',
                    source='combined',
                )
                retries_used += retries
                monthly_rows = monthly_payload.get('data') if isinstance(monthly_payload, dict) else []
                monthly_rows = monthly_rows if isinstance(monthly_rows, list) else []
                with transaction.atomic():
                    for row in monthly_rows:
                        upsert_monthly_record(import_run, supply, row)
                        imported_count += 1
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'monthly',
                    'request_window': {'start': monthly_start.isoformat(), 'end': monthly_end.isoformat()},
                    'attempt_count': retries + 1,
                    'retry_used': retries > 0,
                    'response_code': 200,
                    'account_id': account_id_used,
                    'status': 'success',
                })
            except Exception as exc:  # pylint: disable=broad-except
                failed_ops += 1
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'monthly',
                    'request_window': {'start': monthly_start.isoformat(), 'end': monthly_end.isoformat()},
                    'attempt_count': self.retry_count + 1,
                    'retry_used': self.retry_count > 0,
                    'response_code': None,
                    'status': 'failed',
                    'failure_reason': str(exc),
                })

        invoice_start, invoice_end = get_invoice_window(reporting_month)
        has_invoice_window_data = InvoiceCost.objects.filter(
            supply=supply,
            source_period_end__gte=invoice_start,
            source_period_end__lte=invoice_end,  # inclusive: catches invoices ending exactly on the boundary
        ).exists()
        if not refresh_mode and has_invoice_window_data:
            outcomes.append({
                'supply_id': supply.external_id,
                'data_type': 'invoice',
                'request_window': {'start': invoice_start.isoformat(), 'end': invoice_end.isoformat()},
                'attempt_count': 0,
                'retry_used': False,
                'response_code': None,
                'status': 'skipped',
                'reason': 'cached_data_reused',
            })
        else:
            try:
                invoice_rows, retries, account_id_used = self._fetch_invoices_for_supply(supply=supply)
                retries_used += retries
                skipped_invoice_rows = 0
                with transaction.atomic():
                    for row in invoice_rows:
                        try:
                            upsert_invoice_record(import_run, supply, row)
                            imported_count += 1
                        except ValueError:
                            skipped_invoice_rows += 1
                            logger.warning(
                                "Skipping invoice row with missing datetime fields for supply %s",
                                supply.external_id,
                            )
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'invoice',
                    'request_window': {'start': invoice_start.isoformat(), 'end': invoice_end.isoformat()},
                    'attempt_count': retries + 1,
                    'retry_used': retries > 0,
                    'response_code': 200,
                    'account_id': account_id_used,
                    'skipped_rows': skipped_invoice_rows,
                    'status': 'success',
                })
            except Exception as exc:  # pylint: disable=broad-except
                failed_ops += 1
                outcomes.append({
                    'supply_id': supply.external_id,
                    'data_type': 'invoice',
                    'request_window': {'start': invoice_start.isoformat(), 'end': invoice_end.isoformat()},
                    'attempt_count': self.retry_count + 1,
                    'retry_used': self.retry_count > 0,
                    'response_code': None,
                    'status': 'failed',
                    'failure_reason': str(exc),
                })

        return outcomes, imported_count, retries_used, failed_ops

    def _fetch_with_retry(self, fetcher):
        attempts = 0
        while True:
            try:
                return fetcher(), attempts
            except Exception:  # pylint: disable=broad-except
                if attempts >= self.retry_count:
                    raise
                attempts += 1
                time.sleep(self.retry_backoff)

    def _account_id_candidates(self, supply: Supply) -> List[str]:
        candidates = [
            getattr(settings, 'ETAINABL_ACCOUNT_ID', None),
            supply.external_id,
            supply.parent_account_id,
            supply.name,
        ]
        cleaned = []
        for candidate in candidates:
            if not candidate:
                continue
            value = str(candidate).strip()
            if not value:
                continue
            if value not in cleaned:
                cleaned.append(value)
        return cleaned

    def _fetch_consumption_for_supply(self, supply: Supply, start_date: str, end_date: str, granularity: str, source: str):
        errors = []
        for account_id in self._account_id_candidates(supply):
            try:
                payload, retries = self._fetch_with_retry(
                    lambda: self.client.get_consumption(
                        account_id=account_id,
                        start_date=start_date,
                        end_date=end_date,
                        granularity=granularity,
                        source=source,
                    ),
                )
                return payload, retries, account_id
            except Exception as exc:  # pylint: disable=broad-except
                errors.append(f"{account_id}: {exc}")
        raise Exception('All accountId candidates failed for consumption: ' + ' | '.join(errors))

    def _fetch_invoices_for_supply(self, supply: Supply):
        errors = []
        for account_id in self._account_id_candidates(supply):
            try:
                payload, retries = self._fetch_with_retry(
                    lambda: self.client.get_invoices(
                        account_id=account_id,
                        limit=self.invoice_page_limit,
                        start_page=self.invoice_start_page,
                    ),
                )
                logger.info(
                    "Invoice fetch for supply %s: accountId=%s returned %d rows",
                    supply.external_id, account_id, len(payload) if isinstance(payload, list) else -1,
                )
                return payload, retries, account_id
            except Exception as exc:  # pylint: disable=broad-except
                errors.append(f"{account_id}: {exc}")
        raise Exception('All accountId candidates failed for invoices: ' + ' | '.join(errors))


def get_consumption_display_records(
    reporting_month: str,
    data_type: str = 'monthly',
    supply_external_id: Optional[str] = None,
    supply_external_ids: Optional[List[str]] = None,
) -> List[Dict]:
    model_map = {
        'halfhourly': (HalfHourlyConsumption, 'consumption'),
        'monthly': (MonthlyConsumption, 'consumption'),
        'invoice': (InvoiceCost, 'cost'),
    }
    model, value_field = model_map.get(data_type, model_map['monthly'])

    qs = model.objects.select_related('supply')

    if supply_external_ids:
        qs = qs.filter(supply__external_id__in=supply_external_ids)
    if supply_external_id:
        qs = qs.filter(supply__external_id=supply_external_id)

    # For monthly display, show the trailing 12-month window ending at reporting month.
    if data_type == 'monthly':
        report_start, report_end = reporting_month_bounds(reporting_month)
        window_start = shift_months(report_start, -11)
        qs = qs.filter(source_period_start__gte=window_start, source_period_start__lt=report_end)
    elif data_type == 'invoice':
        window_start, report_end = get_invoice_window(reporting_month)
        window_qs = qs.filter(source_period_end__gte=window_start, source_period_end__lte=report_end)
        if window_qs.exists():
            qs = window_qs
        else:
            # No data in the requested window — fall back to all available invoice history
            # so the user can at least see what has been imported.
            pass  # qs already covers all records for the supply
    else:
        qs = qs.filter(canonical_month_key=reporting_month)

    records = []
    for item in qs.order_by('-source_period_start'):
        records.append({
            'id': item.id,
            'supply_id': item.supply.id,
            'supply_external_id': item.supply.external_id,
            'supply_name': item.supply.name,
            'data_type': data_type,
            'source_period_start': item.source_period_start,
            'source_period_end': item.source_period_end,
            'canonical_month_key': item.canonical_month_key,
            'value': getattr(item, value_field),
            'updated_at': item.updated_at,
        })
    return records


# ============================================================================
# Phase 6: Report Access Scoping Functions
# ============================================================================

def get_reports_for_user(user):
    """
    Get all reports accessible to a user based on their team assignments.
    
    Returns a QuerySet of MonthlyReport objects filtered by:
    - User's team membership (via UserTeamAssignment)
    - User's role in each team
    - Site assignment to teams
    
    If user has no team assignment, returns empty QuerySet.
    If user is admin/superuser, returns all reports.
    
    Args:
        user: The User object for whom to fetch accessible reports
        
    Returns:
        QuerySet: Filtered MonthlyReport objects
    """
    from django.db.models import Q
    from .models import UserTeamAssignment, RoleAssignment
    
    # Admins/superusers see all reports
    if user.is_staff or user.is_superuser:
        return MonthlyReport.objects.all()
    
    # Get user's accessible reports via team assignments
    accessible_reports = get_accessible_reports(user)
    return accessible_reports


def get_accessible_reports(user):
    """
    Get reports accessible to a user with role-based hierarchical access.
    
    Access rules:
    - Admin: All reports
    - Manager: Reports from managed teams + all sub-teams
    - Team Lead: Reports from led team + relevant sub-teams
    - User: Reports from assigned teams only
    - Unassigned: Empty QuerySet
    
    Args:
        user: The User object
        
    Returns:
        QuerySet: Filtered MonthlyReport objects
    """
    from django.db.models import Q
    from .models import RoleAssignment, UserTeamAssignment, Team
    
    # Admins see all reports
    if user.is_staff or user.is_superuser:
        return MonthlyReport.objects.all()
    
    # Build list of accessible teams
    accessible_team_ids = set()
    
    # 1. Teams where user is assigned as a regular member
    user_teams = UserTeamAssignment.objects.filter(
        user=user
    ).values_list('team_id', flat=True)
    accessible_team_ids.update(user_teams)

    # Role-assignment based hierarchical expansion from assigned teams.
    role_names = set(
        RoleAssignment.objects.filter(user=user).values_list('role_name', flat=True)
    )
    if role_names.intersection({'manager', 'team_lead'}):
        for team_id in user_teams:
            team = Team.objects.filter(id=team_id).first()
            if not team:
                continue
            accessible_team_ids.update([sub.id for sub in team.get_sub_teams()])
    
    # 2. Teams where user is the manager (includes all sub-teams)
    for team in Team.objects.filter(manager=user):
        accessible_team_ids.add(team.id)
        accessible_team_ids.update([t.id for t in team.get_sub_teams()])

    # 3. Teams where user is the team lead (includes relevant sub-teams)
    for team in Team.objects.filter(team_lead=user):
        accessible_team_ids.add(team.id)
        accessible_team_ids.update([t.id for t in team.get_sub_teams()])
    
    # If no accessible teams, return empty QuerySet
    if not accessible_team_ids:
        return MonthlyReport.objects.none()
    
    # Team-scoped visibility plus constrained legacy fallback for null-team records.
    # Legacy records remain visible only to directly authorized users.
    return MonthlyReport.objects.filter(
        Q(site__team_id__in=accessible_team_ids)
        |
        (
            Q(site__team_id__isnull=True)
            &
            (
                Q(owner_user=user)
                | Q(created_by_user=user)
                | Q(write_grants__granted_user=user, write_grants__is_active=True)
                | Q(owner_user__team_assignments__team_id__in=accessible_team_ids)
                | Q(created_by_user__team_assignments__team_id__in=accessible_team_ids)
            )
        )
    ).distinct()


# ============================================================================
# Phase 6: Report Access Logging and Caching
# ============================================================================

def log_report_access(user, reports_count, filters_applied=None):
    """
    Log report access event for audit trail.
    
    Args:
        user: The User object accessing reports
        reports_count: Number of reports accessed
        filters_applied: Dict of filters applied (e.g., {'team': 'Engineering', 'status': 'final'})
    """
    filters_str = str(filters_applied) if filters_applied else 'none'
    logger.info(
        'User report access: user=%s, reports_count=%d, filters=%s',
        user.username,
        reports_count,
        filters_str
    )


def get_accessible_reports_cached(user, cache_timeout=300):
    """
    Get accessible reports with caching to reduce database queries.
    
    Caches the result per user session to avoid repeated database hits.
    Cache is keyed by user.id and expires after cache_timeout seconds.
    
    Args:
        user: The User object
        cache_timeout: Cache duration in seconds (default: 5 minutes)
        
    Returns:
        QuerySet: Filtered MonthlyReport objects
    """
    from django.core.cache import cache
    
    cache_key = f'user_accessible_reports_{user.id}'
    cached_reports = cache.get(cache_key)
    
    if cached_reports is not None:
        return cached_reports
    
    # Fetch fresh data if not cached
    reports = get_accessible_reports(user)
    
    # Cache the queryset (converted to list to persist)
    # Note: For large datasets, consider caching just the IDs instead
    report_ids = list(reports.values_list('id', flat=True))
    cache.set(cache_key, report_ids, cache_timeout)
    
    return MonthlyReport.objects.filter(id__in=report_ids)


def invalidate_user_report_cache(user):
    """
    Invalidate cached reports for a user (called when team assignment changes).
    
    Args:
        user: The User object
    """
    from django.core.cache import cache
    
    cache_key = f'user_accessible_reports_{user.id}'
    cache.delete(cache_key)
    logger.info('Invalidated report cache for user=%s', user.username)
