import csv
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from sitesync.models import AuditLogEntry


class AuditLogExportIntegrationTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_user(
            username='parity_admin',
            email='parity_admin@example.com',
            password='StrongPass123!',
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_login(self.admin_user)

        self.matching_one = AuditLogEntry.objects.create(
            actor_user=self.admin_user,
            actor_username_snapshot=self.admin_user.username,
            action_type='REPORT_SAVE_DRAFT',
            action_outcome=AuditLogEntry.OUTCOME_SUCCESS,
            target_entity_type='report',
            target_entity_label='Match A',
            message='Parity match A',
            occurred_at_utc=timezone.now(),
        )
        self.matching_two = AuditLogEntry.objects.create(
            actor_user=self.admin_user,
            actor_username_snapshot=self.admin_user.username,
            action_type='REPORT_SAVE_DRAFT',
            action_outcome=AuditLogEntry.OUTCOME_SUCCESS,
            target_entity_type='report',
            target_entity_label='Match B',
            message='Parity match B',
            occurred_at_utc=timezone.now(),
        )
        AuditLogEntry.objects.create(
            actor_user=self.admin_user,
            actor_username_snapshot=self.admin_user.username,
            action_type='ADMIN_DELETE_USER',
            action_outcome=AuditLogEntry.OUTCOME_SUCCESS,
            target_entity_type='user',
            target_entity_label='Non match',
            message='Should be filtered out',
            occurred_at_utc=timezone.now(),
        )

    def test_export_row_parity_with_filtered_view(self):
        params = {'action_type': 'REPORT_SAVE_DRAFT'}
        viewer_response = self.client.get(reverse('sitesync:admin_audit_logs'), params)
        csv_response = self.client.get(reverse('sitesync:admin_audit_logs_export_csv'), params)
        xlsx_response = self.client.get(reverse('sitesync:admin_audit_logs_export_xlsx'), params)

        self.assertEqual(viewer_response.status_code, 200)
        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)

        csv_rows = list(csv.DictReader(csv_response.content.decode('utf-8').splitlines()))
        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        worksheet = workbook.active
        xlsx_data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))

        # Viewer adds one ADMIN_VIEW_AUDIT_LOG event, but it does not match REPORT_SAVE_DRAFT filter.
        self.assertEqual(len(csv_rows), 2)
        self.assertEqual(len(xlsx_data_rows), 2)
        self.assertContains(viewer_response, 'Matching rows: 2')

    def test_empty_result_export_includes_headers_only(self):
        params = {'action_type': 'NON_EXISTENT_ACTION'}
        csv_response = self.client.get(reverse('sitesync:admin_audit_logs_export_csv'), params)
        xlsx_response = self.client.get(reverse('sitesync:admin_audit_logs_export_xlsx'), params)

        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)

        csv_lines = csv_response.content.decode('utf-8').splitlines()
        self.assertEqual(len(csv_lines), 1)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.max_row, 1)
