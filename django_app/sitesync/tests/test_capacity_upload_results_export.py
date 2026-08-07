from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.db.utils import ProgrammingError
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from sitesync.models import CapacityUploadRun
from sitesync.services import capacity_upload_run_has_row_results


class CapacityUploadResultsExportTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_user(
            username='capacity_export_admin',
            password='StrongPass123!',
            is_staff=True,
        )
        self.regular_user = user_model.objects.create_user(
            username='capacity_export_user',
            password='StrongPass123!',
        )

    def _create_run_with_results(self):
        run = CapacityUploadRun.objects.create(
            uploaded_filename='capacity-results.xlsx',
            total_rows=3,
            accepted_rows=1,
            rejected_rows=2,
            status=CapacityUploadRun.STATUS_PARTIAL_SUCCESS,
            error_summary=[
                'Row 3: Name is blank, eSight Meter Code is blank',
                'Row 4: Av Cap (kVA) must be numeric when provided, Duplicate eSight Meter Code in upload',
            ],
        )
        run.row_results.create(
            source_row_number=2,
            outcome='success',
            explanation='',
            original_columns={
                'Name': 'Meter A',
                'eSight Meter Code': 'MTR-001',
                'Av Cap (kVA)': 25,
                'Region': 'North',
            },
        )
        run.row_results.create(
            source_row_number=3,
            outcome='failure',
            explanation='Name is blank, eSight Meter Code is blank',
            original_columns={
                'Name': '',
                'eSight Meter Code': '',
                'Av Cap (kVA)': 18,
                'Region': 'South',
            },
        )
        run.row_results.create(
            source_row_number=4,
            outcome='failure',
            explanation='Av Cap (kVA) must be numeric when provided, Duplicate eSight Meter Code in upload',
            original_columns={
                'Name': 'Meter B',
                'eSight Meter Code': 'MTR-001',
                'Av Cap (kVA)': 'n/a',
                'Region': 'East',
            },
        )
        return run

    def test_export_requires_authentication(self):
        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        self.assertEqual(response.status_code, 401)

    def test_export_denies_non_admin_user(self):
        self.client.force_login(self.regular_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        self.assertEqual(response.status_code, 403)

    def test_export_returns_xlsx_for_admin(self):
        self._create_run_with_results()
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn('attachment; filename="capacity-upload-results-', response['Content-Disposition'])

    def test_export_workbook_has_successes_and_failures_sheets(self):
        self._create_run_with_results()
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Successes', 'Failures'])

    def test_export_returns_feedback_when_no_completed_run_exists(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'), follow=True)

        self.assertEqual(response.status_code, 200)
        message_text = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any('No completed capacity upload run is available for export.' in message for message in message_text))

    def test_export_returns_feedback_when_latest_run_has_no_row_results(self):
        CapacityUploadRun.objects.create(
            uploaded_filename='capacity-empty.xlsx',
            total_rows=1,
            accepted_rows=0,
            rejected_rows=1,
            status=CapacityUploadRun.STATUS_FAILED,
            error_summary=['Row 2: Name is blank'],
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'), follow=True)

        self.assertEqual(response.status_code, 200)
        message_text = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any('Upload results are unavailable for the latest run.' in message for message in message_text))

    def test_export_row_schema_contains_required_columns(self):
        self._create_run_with_results()
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        workbook = load_workbook(filename=BytesIO(response.content))
        success_sheet = workbook['Successes']
        headers = [cell.value for cell in next(success_sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            headers,
            [
                'Source Row Number',
                'Name',
                'eSight Meter Code',
                'Av Cap (kVA)',
                'Region',
                'Outcome',
                'Explanation',
            ],
        )

    def test_export_combines_multiple_failure_reasons_in_one_explanation_cell(self):
        self._create_run_with_results()
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('sitesync:capacity_upload_results_export'))

        workbook = load_workbook(filename=BytesIO(response.content))
        failure_sheet = workbook['Failures']
        rows = list(failure_sheet.iter_rows(min_row=2, values_only=True))
        target_row = next(row for row in rows if row[0] == 4)
        self.assertEqual(target_row[5], 'failure')
        self.assertIn('Av Cap (kVA) must be numeric when provided', target_row[6])
        self.assertIn('Duplicate eSight Meter Code in upload', target_row[6])

    def test_capacity_upload_row_result_helper_handles_missing_table(self):
        class _BrokenRowResults:
            def exists(self):
                raise ProgrammingError('relation "sitesync_capacityuploadrowresult" does not exist')

        run = SimpleNamespace(row_results=_BrokenRowResults())
        self.assertFalse(capacity_upload_run_has_row_results(run))
