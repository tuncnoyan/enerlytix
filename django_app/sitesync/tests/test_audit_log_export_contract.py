import csv
from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from sitesync.models import AuditLogEntry


class AuditLogExportContractTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_user(
            username='export_admin',
            email='export_admin@example.com',
            password='StrongPass123!',
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_login(self.admin_user)
        AuditLogEntry.objects.create(
            actor_user=self.admin_user,
            actor_username_snapshot=self.admin_user.username,
            action_type='REPORT_SAVE_DRAFT',
            action_outcome=AuditLogEntry.OUTCOME_SUCCESS,
            target_entity_type='report',
            target_entity_label='Export report',
            message='Saved report draft for export tests.',
            occurred_at_utc=timezone.now(),
        )

    def test_csv_export_content_type_and_headers(self):
        response = self.client.get(reverse('sitesync:admin_audit_logs_export_csv'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        decoded = response.content.decode('utf-8')
        self.assertIn('utc_timestamp', decoded.splitlines()[0])

    def test_xlsx_export_content_type_and_headers(self):
        response = self.client.get(reverse('sitesync:admin_audit_logs_export_xlsx'))

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[0], 'utc_timestamp')

    def test_invalid_filters_return_400_for_exports(self):
        start = timezone.now()
        end = start - timedelta(days=2)

        csv_response = self.client.get(
            reverse('sitesync:admin_audit_logs_export_csv'),
            {'start': start.isoformat(), 'end': end.isoformat()},
        )
        xlsx_response = self.client.get(
            reverse('sitesync:admin_audit_logs_export_xlsx'),
            {'start': start.isoformat(), 'end': end.isoformat()},
        )

        self.assertEqual(csv_response.status_code, 400)
        self.assertEqual(xlsx_response.status_code, 400)

    def test_timestamps_are_exported_with_utc_label(self):
        response = self.client.get(reverse('sitesync:admin_audit_logs_export_csv'))
        rows = list(csv.DictReader(response.content.decode('utf-8').splitlines()))

        self.assertGreaterEqual(len(rows), 1)
        self.assertTrue(rows[0]['utc_timestamp'].endswith('UTC'))

    def test_export_threshold_guard_fails_fast_over_50000_rows(self):
        now = timezone.now()
        bulk_rows = [
            AuditLogEntry(
                actor_user=self.admin_user,
                actor_username_snapshot=self.admin_user.username,
                action_type='REPORT_SAVE_DRAFT',
                action_outcome=AuditLogEntry.OUTCOME_SUCCESS,
                target_entity_type='report',
                target_entity_label='Threshold row',
                message='Threshold data row',
                occurred_at_utc=now,
            )
            for _ in range(50001)
        ]
        AuditLogEntry.objects.bulk_create(bulk_rows, batch_size=2000)

        response = self.client.get(reverse('sitesync:admin_audit_logs_export_csv'))

        self.assertEqual(response.status_code, 400)
        self.assertIn('narrow filters', response.json().get('detail', '').lower())
